"""
Sync Excel Ledger (Branch 2) → Database
Syncs classes, students, accounts, fees and payments from '2. Ledger - Feb 26.xlsx'.

Usage:
    python sync_excel_br2.py              # dry-run (shows what would happen)
    python sync_excel_br2.py --apply      # actually writes to DB
"""

import os
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from students.models import Class, Student
from schools.models import School, Organization
from finance.models import Account, FeeStructure, FeePayment
from decimal import Decimal
from datetime import date
from openpyxl import load_workbook

# --- Configuration ---

SCHOOL_ID = 2
ORG_ID = 1
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '2. Ledger - Feb 26.xlsx')
INCOME_SHEET = 'Income'
CONTACT_SHEET = 'Br2 Income Exp - AugSep 25'

# Excel column indices for Income sheet (1-based)
COL_BRANCH = 1    # A
COL_CLASS = 2     # B
COL_ROLL = 3      # C
COL_NAME = 4      # D
COL_FEE = 5       # E
COL_TOTAL_PAYABLE = 6  # F
COL_RECEIVED = 7       # G
COL_ACCOUNT = 8        # H
COL_BALANCE = 9        # I

BRANCH_2 = 'Branch 2'

# Excel class name → DB class name
CLASS_MAP = {
    'PG': 'Playgroup',
    'J1': 'Junior 1',
    'J2': 'Junior 2',
    'Cl1A': 'Class 1',
    'Cl2': 'Class 2',
    'Cl3': 'Class 3',
    'Cl4': 'Class 4',
    'Cl5': 'Class 5',
}

# Branch 2 school-level accounts
ACCOUNT_MAP = {
    'Principal Br2': {'db_name': 'Principal', 'type': 'CASH', 'bbf': Decimal('22220'), 'staff_visible': True},
    'Fund Br2':      {'db_name': 'Fund',      'type': 'CASH', 'bbf': Decimal('11210'), 'staff_visible': True},
}

# Shared person accounts (exist on school=1, org=1)
SHARED_ACCOUNTS = ['Shah Mir', 'Abdul Abbas', 'Agha Guloon']

# Contact sheet class sections: (class_code, header_row, data_start_row, data_end_row)
CONTACT_SECTIONS = [
    ('PG',   2,   4,  37),
    ('J1',  39,  41,  68),
    ('J2',  72,  74,  94),
    ('Cl1A', 96,  98, 113),
    ('Cl2', 115, 117, 132),
    ('Cl3', 134, 136, 144),
    ('Cl4', 146, 148, 156),
    ('Cl5', 158, 160, 171),
]


def read_excel_students():
    """Read Branch 2 students from Income sheet, grouped by class."""
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[INCOME_SHEET]

    students_by_class = {}
    current_class = None
    in_branch_2 = False

    for row in ws.iter_rows(min_row=2, max_col=10, values_only=False):
        branch = row[COL_BRANCH - 1].value
        cls_name = row[COL_CLASS - 1].value
        roll = row[COL_ROLL - 1].value
        name = row[COL_NAME - 1].value
        fee = row[COL_FEE - 1].value
        total_payable = row[COL_TOTAL_PAYABLE - 1].value
        received = row[COL_RECEIVED - 1].value
        account = row[COL_ACCOUNT - 1].value

        # Start reading when we hit Branch 2
        if branch and str(branch).strip() == BRANCH_2:
            in_branch_2 = True

        if not in_branch_2:
            continue

        # Track current class
        if cls_name and str(cls_name).strip() in CLASS_MAP:
            current_class = str(cls_name).strip()

        # Skip rows without a student name or roll number
        if not name or not roll:
            continue

        # Skip non-student rows (like "Copies" or totals)
        name_str = str(name).strip()
        if name_str.lower() in ('total', 'copies', ''):
            continue

        if current_class and current_class in CLASS_MAP:
            if current_class not in students_by_class:
                students_by_class[current_class] = []

            roll_str = str(int(roll)) if isinstance(roll, (int, float)) else str(roll).strip()

            # Check for duplicate roll within the same class
            existing_rolls = {s['roll'] for s in students_by_class[current_class]}
            if roll_str in existing_rolls:
                max_roll = max(int(s['roll']) for s in students_by_class[current_class])
                roll_str = str(max_roll + 1)
                print(f'  [WARN] Duplicate roll in Excel for {current_class}: "{name_str}" reassigned to Roll #{roll_str}')

            monthly_fee = Decimal(str(fee)) if fee and fee != 0 else Decimal('0')
            payable = Decimal(str(total_payable)) if total_payable and total_payable != 0 else monthly_fee
            received_amt = Decimal(str(received)) if received and received != 0 else Decimal('0')
            acct_name = str(account).strip() if account else None

            students_by_class[current_class].append({
                'roll': roll_str,
                'name': name_str,
                'monthly_fee': monthly_fee,
                'total_payable': payable,
                'received': received_amt,
                'previous_balance': payable - monthly_fee,
                'account_name': acct_name,
            })

    wb.close()

    # Summary
    total = sum(len(v) for v in students_by_class.values())
    print(f'\n  Read {total} Branch 2 students from Income sheet:')
    for cls, students in students_by_class.items():
        print(f'    {CLASS_MAP[cls]} ({cls}): {len(students)} students')

    return students_by_class


def read_contact_info():
    """Read father name, phone, whatsapp from Br2 Income Exp sheet."""
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[CONTACT_SHEET]

    contacts = {}  # (class_code, roll) -> {father, phone, whatsapp}

    for cls_code, header_row, start_row, end_row in CONTACT_SECTIONS:
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=6, values_only=False):
            roll_val = row[0].value  # A = Roll No
            name_val = row[1].value  # B = Name
            father = row[2].value    # C = Father's Name
            phone = row[3].value     # D = Phone No
            whatsapp = row[4].value  # E = WhatsApp Number

            if roll_val is None or name_val is None:
                continue
            if not isinstance(roll_val, (int, float)):
                continue

            roll_str = str(int(roll_val))
            contacts[(cls_code, roll_str)] = {
                'father': str(father).strip() if father else '',
                'phone': str(phone).strip() if phone else '',
                'whatsapp': str(whatsapp).strip() if whatsapp else '',
            }

    wb.close()
    print(f'\n  Read {len(contacts)} contact records from Br2 Income Exp sheet')
    return contacts


def sync_accounts(apply=False):
    """Create/update Branch 2 school-level accounts."""
    school = School.objects.get(id=SCHOOL_ID)

    print('\n=== ACCOUNT SYNC ===')
    for excel_name, cfg in ACCOUNT_MAP.items():
        db_name = cfg['db_name']
        try:
            acct = Account.objects.get(school=school, name=db_name)
            old_bbf = acct.opening_balance
            if old_bbf != cfg['bbf']:
                if apply:
                    acct.opening_balance = cfg['bbf']
                    acct.staff_visible = cfg['staff_visible']
                    acct.save()
                    print(f'  {db_name}: UPDATED BBF {old_bbf} -> {cfg["bbf"]}')
                else:
                    print(f'  {db_name}: EXISTS (ID={acct.id}) BBF={old_bbf} -> will update to {cfg["bbf"]}')
            else:
                print(f'  {db_name}: OK (ID={acct.id}, BBF={acct.opening_balance})')
        except Account.DoesNotExist:
            if apply:
                acct = Account.objects.create(
                    school=school,
                    organization_id=ORG_ID,
                    name=db_name,
                    account_type=cfg['type'],
                    opening_balance=cfg['bbf'],
                    is_active=True,
                    staff_visible=cfg['staff_visible'],
                )
                print(f'  {db_name}: CREATED (ID={acct.id}, type={cfg["type"]}, BBF={cfg["bbf"]})')
            else:
                print(f'  {db_name}: MISSING -> will create (type={cfg["type"]}, BBF={cfg["bbf"]})')

    # Verify shared accounts are accessible
    print('\n  Shared accounts (org-level):')
    for name in SHARED_ACCOUNTS:
        try:
            acct = Account.objects.get(name=name, organization_id=ORG_ID)
            print(f'    {name}: OK (ID={acct.id}, school={acct.school_id})')
        except Account.DoesNotExist:
            print(f'    {name}: NOT FOUND — payments to this account will have no account link')


def sync_classes_and_students(apply=False):
    """Create classes and students for Branch 2, enriched with contact info."""
    school = School.objects.get(id=SCHOOL_ID)
    excel_data = read_excel_students()
    contacts = read_contact_info()

    total_new = 0
    total_match = 0
    total_mismatch = 0
    total_db_only = 0
    total_enriched = 0

    print(f'\nSchool: {school.name} (ID={school.id})')
    print('=' * 60)

    # --- Class Sync ---
    print('\n=== CLASS SYNC ===')
    db_classes = {}
    for excel_cls, db_cls_name in CLASS_MAP.items():
        try:
            cls = Class.objects.get(school=school, name=db_cls_name)
            db_classes[excel_cls] = cls
            print(f'  {excel_cls} -> {db_cls_name}: EXISTS (ID={cls.id})')
        except Class.DoesNotExist:
            if apply:
                cls = Class.objects.create(school=school, name=db_cls_name, is_active=True)
                db_classes[excel_cls] = cls
                print(f'  {excel_cls} -> {db_cls_name}: CREATED (ID={cls.id})')
            else:
                print(f'  {excel_cls} -> {db_cls_name}: MISSING — will create')

    # --- Student Sync ---
    print('\n=== STUDENT SYNC ===')
    for excel_cls in CLASS_MAP:
        db_cls_name = CLASS_MAP[excel_cls]
        cls = db_classes.get(excel_cls)
        if not cls:
            print(f'\n{db_cls_name} ({excel_cls}): SKIPPED (class not found — run with --apply first)')
            continue

        excel_students = excel_data.get(excel_cls, [])
        db_students = list(Student.objects.filter(school=school, class_obj=cls, is_active=True))

        # Build lookup by roll number
        db_by_roll = {s.roll_number: s for s in db_students}
        excel_rolls = {s['roll'] for s in excel_students}

        class_new = 0
        class_match = 0
        class_mismatch = 0

        print(f'\n{db_cls_name} ({excel_cls}): {len(excel_students)} in Excel, {len(db_students)} in DB')

        for es in excel_students:
            roll = es['roll']
            name = es['name']

            # Get contact info
            contact = contacts.get((excel_cls, roll), {})

            if roll in db_by_roll:
                db_s = db_by_roll[roll]
                if db_s.name.strip().lower() == name.strip().lower():
                    class_match += 1
                    # Enrich with contact info if available
                    if contact and apply:
                        updated = False
                        if contact.get('father') and not db_s.parent_name:
                            db_s.parent_name = contact['father']
                            updated = True
                        if contact.get('phone') and not db_s.parent_phone:
                            db_s.parent_phone = contact['phone']
                            updated = True
                        if updated:
                            db_s.save()
                            total_enriched += 1
                else:
                    print(f'  [MISMATCH] Roll #{roll}: Excel="{name}" vs DB="{db_s.name}"')
                    class_mismatch += 1
            else:
                if apply:
                    _, created = Student.objects.get_or_create(
                        school=school,
                        class_obj=cls,
                        roll_number=roll,
                        defaults={
                            'name': name,
                            'is_active': True,
                            'parent_name': contact.get('father', ''),
                            'parent_phone': contact.get('phone', ''),
                        },
                    )
                    if created:
                        print(f'  [CREATED] Roll #{roll} {name}')
                        if contact.get('father'):
                            total_enriched += 1
                    else:
                        print(f'  [EXISTS] Roll #{roll} {name} — already in DB (skipped)')
                else:
                    father_info = f' (Father: {contact["father"]})' if contact.get('father') else ''
                    print(f'  [NEW] Roll #{roll} {name}{father_info} — will create')
                class_new += 1

        # Students in DB but not in Excel
        for roll, db_s in db_by_roll.items():
            if roll not in excel_rolls:
                print(f'  [DB ONLY] Roll #{roll} {db_s.name} — in DB but not in Excel')
                total_db_only += 1

        total_new += class_new
        total_match += class_match
        total_mismatch += class_mismatch

        if class_new == 0 and class_mismatch == 0:
            print(f'  All {class_match} students matched.')

    # --- Summary ---
    print('\n' + '=' * 60)
    print('=== SUMMARY ===')
    print(f'  Matches:    {total_match}')
    print(f'  New:        {total_new}')
    print(f'  Mismatches: {total_mismatch}')
    print(f'  DB only:    {total_db_only}')
    print(f'  Enriched:   {total_enriched} (contact info added)')

    if not apply and total_new > 0:
        print(f'\n  Run with --apply to create {total_new} new students.')
    elif apply and total_new > 0:
        print(f'\n  {total_new} students created successfully.')
    else:
        print('\n  Nothing to do — all synced.')


def sync_fee_structures(apply=False):
    """Delete old fee data for Branch 2, create student-level FeeStructures from Excel."""
    school = School.objects.get(id=SCHOOL_ID)
    excel_data = read_excel_students()

    # --- Delete existing fee data for Branch 2 ---
    existing_payments = FeePayment.objects.filter(school=school).count()
    existing_structures = FeeStructure.objects.filter(school=school).count()

    print('\n=== FEE STRUCTURE SYNC ===')
    print(f'  Existing FeePayments (Branch 2): {existing_payments} -> will DELETE')
    print(f'  Existing FeeStructures (Branch 2): {existing_structures} -> will DELETE')

    if apply:
        FeePayment.objects.filter(school=school).delete()
        FeeStructure.objects.filter(school=school).delete()
        print('  Deleted.')

    # --- Create student-level FeeStructures ---
    created = 0
    skipped = 0
    effective_from = date(2026, 2, 1)

    for excel_cls, db_cls_name in CLASS_MAP.items():
        students = excel_data.get(excel_cls, [])
        try:
            cls = Class.objects.get(school=school, name=db_cls_name)
        except Class.DoesNotExist:
            print(f'  {db_cls_name}: class not found, skipping')
            continue

        class_created = 0
        for es in students:
            roll = es['roll']
            fee = es['monthly_fee']

            try:
                student = Student.objects.get(school=school, class_obj=cls, roll_number=roll)
            except Student.DoesNotExist:
                skipped += 1
                continue

            if apply:
                FeeStructure.objects.get_or_create(
                    school=school,
                    student=student,
                    class_obj=None,
                    is_active=True,
                    defaults={
                        'monthly_amount': fee,
                        'effective_from': effective_from,
                    },
                )
            class_created += 1

        created += class_created
        print(f'  {db_cls_name}: {class_created} student fee structures {"created" if apply else "to create"}')

    print(f'\n  Total: {created} structures, {skipped} skipped (student not in DB)')


def sync_fee_payments(apply=False):
    """Create FeePayment records with received amounts from Excel."""
    school = School.objects.get(id=SCHOOL_ID)
    excel_data = read_excel_students()

    # Build account lookup
    db_accounts = {}

    # School-level accounts (Branch 2)
    for excel_name, cfg in ACCOUNT_MAP.items():
        try:
            db_accounts[excel_name] = Account.objects.get(school=school, name=cfg['db_name'])
        except Account.DoesNotExist:
            print(f'  [WARN] Account "{cfg["db_name"]}" not found for Branch 2')

    # Shared accounts (org-level, on school=1)
    for shared_name in SHARED_ACCOUNTS:
        try:
            db_accounts[shared_name] = Account.objects.get(name=shared_name, organization_id=ORG_ID)
        except Account.DoesNotExist:
            print(f'  [WARN] Shared account "{shared_name}" not found')

    # Build flat payment list
    payments = []
    for excel_cls, students in excel_data.items():
        for s in students:
            payments.append({
                'class': excel_cls,
                'roll': s['roll'],
                'name': s['name'],
                'monthly_fee': s['monthly_fee'],
                'total_payable': s['total_payable'],
                'received': s['received'],
                'previous_balance': s['previous_balance'],
                'account_name': s['account_name'],
            })

    print('\n=== FEE PAYMENT SYNC (Feb 2026) ===')
    updated = 0
    created = 0
    skipped = 0
    not_found = 0
    total_received = Decimal('0')
    account_totals = {}

    for p in payments:
        db_cls_name = CLASS_MAP.get(p['class'])
        if not db_cls_name:
            continue

        try:
            cls = Class.objects.get(school=school, name=db_cls_name)
            student = Student.objects.get(school=school, class_obj=cls, roll_number=p['roll'])
        except (Class.DoesNotExist, Student.DoesNotExist):
            if p['received'] > 0:
                print(f'  [NOT FOUND] {db_cls_name} Roll#{p["roll"]} {p["name"]}: student not in DB')
                not_found += 1
            continue

        # Resolve account
        db_account = db_accounts.get(p['account_name']) if p['account_name'] else None

        # Find or create FeePayment
        try:
            fp = FeePayment.objects.get(school=school, student=student, month=2, year=2026)
        except FeePayment.DoesNotExist:
            if apply:
                fp = FeePayment(
                    school=school,
                    student=student,
                    month=2,
                    year=2026,
                    amount_due=p['total_payable'] if p['total_payable'] > 0 else p['monthly_fee'],
                    previous_balance=p['previous_balance'],
                    amount_paid=p['received'],
                )
                if db_account:
                    fp.account = db_account
                fp.save()
            total_received += p['received']
            if p['account_name']:
                account_totals[p['account_name']] = account_totals.get(p['account_name'], Decimal('0')) + p['received']
            print(f'  {"[CREATED]" if apply else "[WILL CREATE]"} {db_cls_name} Roll#{p["roll"]} {p["name"]}: due={p["total_payable"]}, paid={p["received"]}, account={p["account_name"]}')
            created += 1
            continue

        # Check if update needed
        needs_update = False
        changes = []

        if fp.amount_due != p['total_payable'] and p['total_payable'] > 0:
            changes.append(f'due: {fp.amount_due}->{p["total_payable"]}')
            needs_update = True

        if fp.amount_paid != p['received']:
            changes.append(f'paid: {fp.amount_paid}->{p["received"]}')
            needs_update = True

        if db_account and fp.account_id != db_account.id:
            changes.append(f'account: {db_account.name}')
            needs_update = True

        if not needs_update:
            skipped += 1
            continue

        if apply:
            if p['total_payable'] > 0:
                fp.previous_balance = p['previous_balance']
                fp.amount_due = p['total_payable']
            fp.amount_paid = p['received']
            if db_account:
                fp.account = db_account
            fp.save()

        total_received += p['received']
        if p['account_name']:
            account_totals[p['account_name']] = account_totals.get(p['account_name'], Decimal('0')) + p['received']

        if changes:
            print(f'  {"[UPDATED]" if apply else "[WILL UPDATE]"} {db_cls_name} Roll#{p["roll"]} {p["name"]}: {", ".join(changes)}')
        updated += 1

    # Summary
    print(f'\n  Created: {created}')
    print(f'  Updated: {updated}')
    print(f'  Skipped (no change): {skipped}')
    print(f'  Not found: {not_found}')
    print(f'  Total received: {total_received:,.0f}')
    for acct, total in sorted(account_totals.items()):
        print(f'    {acct}: {total:,.0f}')


if __name__ == '__main__':
    apply = '--apply' in sys.argv
    if apply:
        print('MODE: APPLY (will write to database)')
    else:
        print('MODE: DRY RUN (no changes will be made)')
    sync_accounts(apply=apply)
    sync_classes_and_students(apply=apply)
    sync_fee_structures(apply=apply)
    sync_fee_payments(apply=apply)
