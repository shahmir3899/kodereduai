import io
import re
from decimal import Decimal
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from pgvector.django import CosineDistance
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from core.mixins import TenantQuerySetMixin, ensure_tenant_school_id
from core.permissions import ADMIN_ROLES, IsSchoolAdmin, IsSchoolAdminOrReadOnly, HasSchoolAccess, ModuleAccessMixin, CanManageStudentAssessments, get_effective_role, get_teacher_combined_scope
from core.class_scope import resolve_class_scope
from core.ai_jobs import complete_ai_job, create_ai_job, fail_ai_job
from core.embeddings import generate_text_embedding
from lms.models import Tag, QuestionTag

from .models import (
    ExamType, ExamGroup, Exam, ExamSubject, StudentMark, GradeScale,
    Question, ExamPaper, PaperQuestion, StudentResponse, PaperUpload, PaperFeedback,
    StudentTermAssessment,
)
from .serializers import (
    ExamTypeSerializer, ExamTypeCreateSerializer,
    ExamSerializer, ExamCreateSerializer,
    BulkTestRequestSerializer,
    ExamSubjectSerializer, ExamSubjectCreateSerializer,
    StudentMarkSerializer, StudentMarkCreateSerializer,
    StudentMarkBulkEntrySerializer,
    StudentResponseSerializer, StudentResponseBulkSubmitSerializer,
    GradeScaleSerializer, GradeScaleCreateSerializer,
    ExamGroupSerializer, ExamGroupCreateSerializer,
    ExamGroupWizardCreateSerializer, DateSheetUpdateSerializer,
    QuestionSerializer, QuestionCreateUpdateSerializer,
    ExamPaperSerializer, ExamPaperCreateUpdateSerializer,
    ExamPaperDraftEnsureSerializer, ExamPaperDraftAutosaveSerializer,
    PaperUploadSerializer, PaperUploadCreateSerializer,
    StudentTermAssessmentSerializer,
    PaperFeedbackSerializer, QuestionReviewSerializer,
)
from .tasks import recompute_question_stats



def _resolve_school_id(request):
    school_id = ensure_tenant_school_id(request)
    if school_id:
        return school_id
    # If X-School-ID header was sent but rejected, don't fall back
    if request.headers.get('X-School-ID'):
        return None
    sid = (
        request.query_params.get('school_id')
        or request.data.get('school_id')
        or request.data.get('school')
    )
    if sid:
        return int(sid)
    if request.user.school_id:
        return request.user.school_id
    return None


def _apply_teacher_exam_scope(qs, request, class_field='class_obj_id', subject_field=None):
    """Apply dual-layer teacher scope for exam-related querysets.
    Uses section-class scope when teacher has session assignments."""
    if get_effective_role(request) != 'TEACHER':
        return qs

    school_id = _resolve_school_id(request)
    scope = get_teacher_combined_scope(request, school_id=school_id)
    all_class_ids = scope['all_class_ids']
    full_class_ids = scope['full_class_ids']
    session_ids = scope.get('full_session_class_ids', set())

    # For class-level entities (Exam), visibility is union of full class scope and subject assignment classes.
    if not subject_field:
        if all_class_ids:
            return qs.filter(**{f'{class_field}__in': all_class_ids})
        return qs.none()

    predicates = Q()

    if session_ids:
        # Section-scoped: teacher's class-teacher assignment is section-level
        # Access exams only for those specific master class IDs
        if full_class_ids:
            predicates |= Q(**{f'{class_field}__in': full_class_ids})
    elif full_class_ids:
        predicates |= Q(**{f'{class_field}__in': full_class_ids})

    for class_id, subject_ids in scope['class_subject_map'].items():
        if subject_ids:
            predicates |= Q(**{class_field: class_id, f'{subject_field}__in': list(subject_ids)})

    if not predicates:
        return qs.none()
    return qs.filter(predicates)


def _get_teacher_class_subject_map(request, school_id=None):
    """Return {class_id: set(subject_ids)} for teacher-assigned subject scope."""
    if get_effective_role(request) != 'TEACHER':
        return {}
    school_id = school_id or _resolve_school_id(request)
    scope = get_teacher_combined_scope(request, school_id=school_id)
    return scope.get('class_subject_map', {})


def _get_teacher_allowed_subject_ids(request, school_id=None):
    """Flatten teacher class-subject assignments into a subject ID set."""
    class_subject_map = _get_teacher_class_subject_map(request, school_id=school_id)
    allowed = set()
    for subject_ids in class_subject_map.values():
        allowed.update(subject_ids)
    return allowed


def _is_teacher_allowed_for_class_subject(request, class_id, subject_id, school_id=None):
    """Check strict teacher assignment for a class-subject pair."""
    if get_effective_role(request) != 'TEACHER':
        return True
    if class_id is None or subject_id is None:
        return False
    class_subject_map = _get_teacher_class_subject_map(request, school_id=school_id)
    try:
        class_id = int(class_id)
        subject_id = int(subject_id)
    except (TypeError, ValueError):
        return False
    return subject_id in class_subject_map.get(class_id, set())


def _is_teacher_class_teacher_for_class(request, class_id, school_id=None):
    """Check whether current teacher has class-teacher scope for this class."""
    if get_effective_role(request) != 'TEACHER':
        return False
    try:
        class_id = int(class_id)
    except (TypeError, ValueError):
        return False
    scope = get_teacher_combined_scope(request, school_id=school_id)
    return class_id in scope.get('full_class_ids', set())


def _can_manage_exam_papers(request, class_id=None, subject_id=None, school_id=None):
    """Return True when role is allowed to create/update exam papers."""
    role = get_effective_role(request)
    if role in ADMIN_ROLES:
        return True
    if role != 'TEACHER':
        return False
    return _is_teacher_class_teacher_for_class(request, class_id, school_id=school_id)


def _short_academic_year_name(name):
    return re.sub(r'^academic\s+year\s*', '', name or '', flags=re.IGNORECASE).strip()


def _generate_bulk_test_name(subject_name, term_name=None, academic_year_name=None):
    name = f'Test - {subject_name}'
    if term_name and term_name.lower() not in name.lower():
        name += f' - {term_name}'
    year_short = _short_academic_year_name(academic_year_name)
    if year_short:
        name += f' {year_short}'
    return name


def _assert_bulk_test_role(request):
    role = get_effective_role(request)
    if role not in ADMIN_ROLES and role != 'TEACHER':
        raise PermissionDenied('You do not have permission to manage tests.')


def _build_bulk_test_plan(request, data):
    school_id = _resolve_school_id(request)
    if not school_id:
        raise ValidationError({'detail': 'No school context.'})

    _assert_bulk_test_role(request)

    from academic_sessions.models import AcademicYear, Term
    from academics.models import ClassSubject, Subject
    from students.models import Class

    academic_year = AcademicYear.objects.filter(
        school_id=school_id,
        id=data['academic_year'],
        is_active=True,
    ).only('id', 'name').first()
    if not academic_year:
        raise ValidationError({'academic_year': 'Academic year is invalid for the active school.'})

    term = None
    if data.get('term'):
        term = Term.objects.filter(
            school_id=school_id,
            academic_year_id=academic_year.id,
            id=data['term'],
        ).only('id', 'name').first()
        if not term:
            raise ValidationError({'term': 'Term is invalid for the selected academic year.'})

    exam_type = ExamType.objects.filter(
        school_id=school_id,
        id=data['exam_type'],
        is_active=True,
    ).only('id', 'name').first()
    if not exam_type:
        raise ValidationError({'exam_type': 'Exam type is invalid for the active school.'})

    class_obj = Class.objects.filter(
        school_id=school_id,
        id=data['class_obj'],
        is_active=True,
    ).only('id', 'name').first()
    if not class_obj:
        raise ValidationError({'class_obj': 'Class is invalid for the active school.'})

    role = get_effective_role(request)
    teacher_scope = get_teacher_combined_scope(
        request,
        school_id=school_id,
        academic_year_id=academic_year.id,
    )
    if role == 'TEACHER' and class_obj.id not in teacher_scope['all_class_ids']:
        raise PermissionDenied('You do not have access to this class.')

    requested_subject_ids = [row['subject_id'] for row in data['tests']]
    subjects_by_id = {
        subject.id: subject
        for subject in Subject.objects.filter(
            school_id=school_id,
            id__in=requested_subject_ids,
            is_active=True,
        ).only('id', 'name', 'code')
    }

    class_subjects_qs = ClassSubject.objects.filter(
        school_id=school_id,
        class_obj_id=class_obj.id,
        subject_id__in=requested_subject_ids,
        is_active=True,
    ).filter(
        Q(academic_year_id=academic_year.id) | Q(academic_year__isnull=True)
    ).select_related('subject')

    if role == 'TEACHER':
        class_subjects_qs = class_subjects_qs.filter(teacher__user=request.user)

    assigned_subjects = {}
    for class_subject in class_subjects_qs.order_by('-academic_year_id', '-id'):
        assigned_subjects.setdefault(class_subject.subject_id, class_subject)

    existing_tests_qs = ExamSubject.objects.filter(
        school_id=school_id,
        subject_id__in=requested_subject_ids,
        is_active=True,
        exam__school_id=school_id,
        exam__academic_year_id=academic_year.id,
        exam__class_obj_id=class_obj.id,
        exam__exam_type_id=exam_type.id,
        exam__exam_group__isnull=True,
        exam__is_active=True,
    ).select_related('exam', 'subject')
    if term:
        existing_tests_qs = existing_tests_qs.filter(exam__term_id=term.id)
    else:
        existing_tests_qs = existing_tests_qs.filter(exam__term__isnull=True)

    existing_by_subject = {}
    for exam_subject in existing_tests_qs:
        existing_by_subject.setdefault(exam_subject.subject_id, exam_subject)

    items = []
    for row in data['tests']:
        subject = subjects_by_id.get(row['subject_id'])
        class_subject = assigned_subjects.get(row['subject_id'])
        existing_exam_subject = existing_by_subject.get(row['subject_id'])

        if not subject:
            items.append({
                'subject_id': row['subject_id'],
                'subject_name': None,
                'subject_code': None,
                'name': row.get('name', '').strip(),
                'exam_date': row['exam_date'],
                'total_marks': row['total_marks'],
                'start_time': row.get('start_time'),
                'end_time': row.get('end_time'),
                'status': 'invalid',
                'reason': 'Subject is invalid for the active school.',
            })
            continue

        resolved_name = row.get('name', '').strip() or _generate_bulk_test_name(
            subject.name,
            term_name=term.name if term else None,
            academic_year_name=academic_year.name,
        )

        if not class_subject:
            reason = 'You are not assigned to this class-subject pair.' if role == 'TEACHER' else 'Subject is not assigned to the selected class.'
            items.append({
                'subject_id': subject.id,
                'subject_name': subject.name,
                'subject_code': subject.code,
                'name': resolved_name,
                'exam_date': row['exam_date'],
                'total_marks': row['total_marks'],
                'start_time': row.get('start_time'),
                'end_time': row.get('end_time'),
                'status': 'forbidden' if role == 'TEACHER' else 'invalid',
                'reason': reason,
            })
            continue

        if existing_exam_subject:
            items.append({
                'subject_id': subject.id,
                'subject_name': subject.name,
                'subject_code': subject.code,
                'name': resolved_name,
                'exam_date': row['exam_date'],
                'total_marks': row['total_marks'],
                'start_time': row.get('start_time'),
                'end_time': row.get('end_time'),
                'status': 'conflict',
                'reason': f'Active test "{existing_exam_subject.exam.name}" already exists for this subject.',
                'existing_exam_id': existing_exam_subject.exam_id,
            })
            continue

        items.append({
            'subject_id': subject.id,
            'subject_name': subject.name,
            'subject_code': subject.code,
            'name': resolved_name,
            'exam_date': row['exam_date'],
            'total_marks': row['total_marks'],
            'start_time': row.get('start_time'),
            'end_time': row.get('end_time'),
            'status': 'create',
            'reason': '',
        })

    counts = {
        'requested': len(items),
        'create': sum(1 for item in items if item['status'] == 'create'),
        'conflict': sum(1 for item in items if item['status'] == 'conflict'),
        'forbidden': sum(1 for item in items if item['status'] == 'forbidden'),
        'invalid': sum(1 for item in items if item['status'] == 'invalid'),
    }

    return {
        'class_obj': class_obj.id,
        'class_name': class_obj.name,
        'academic_year': academic_year.id,
        'academic_year_name': academic_year.name,
        'term': term.id if term else None,
        'term_name': term.name if term else None,
        'exam_type': exam_type.id,
        'exam_type_name': exam_type.name,
        'counts': counts,
        'can_apply': counts['create'] > 0 and counts['conflict'] == 0 and counts['forbidden'] == 0 and counts['invalid'] == 0,
        'tests': items,
    }


class ExamTypeViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    required_module = 'examinations'
    queryset = ExamType.objects.all()
    permission_classes = [IsAuthenticated, IsSchoolAdminOrReadOnly, HasSchoolAccess]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ExamTypeCreateSerializer
        return ExamTypeSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['school_id'] = _resolve_school_id(self.request)
        return ctx

    def get_queryset(self):
        return super().get_queryset()


def _build_date_sheet_grid(group, school_id):
    """Pivot an ExamGroup's per-(class, subject) exam dates into a Date x Class grid.

    Subjects with no exam_date yet can't be placed on a date-indexed grid, so
    they're returned separately as 'unscheduled' rather than silently dropped.
    Two subjects landing on the same date for the same class (not prevented
    anywhere upstream) are joined into one cell rather than treated as an error.
    """
    exam_subjects = ExamSubject.objects.filter(
        exam__exam_group=group, exam__is_active=True,
        is_active=True, school_id=school_id,
    ).select_related('subject', 'exam', 'exam__class_obj').order_by(
        'exam__class_obj__grade_level', 'exam__class_obj__name', 'subject__name',
    )

    columns_by_id = {}
    cells = {}  # (date, class_id) -> [subject_name, ...]
    unscheduled = []

    for es in exam_subjects:
        cls = es.exam.class_obj
        if cls.id not in columns_by_id:
            columns_by_id[cls.id] = {
                'class_id': cls.id,
                'label': f'{cls.name} - {cls.section}' if cls.section else cls.name,
                '_sort_key': (cls.grade_level if cls.grade_level is not None else 0, cls.name),
            }
        if not es.exam_date:
            unscheduled.append({
                'subject_name': es.subject.name,
                'class_name': columns_by_id[cls.id]['label'],
            })
            continue
        cells.setdefault((es.exam_date, cls.id), []).append(es.subject.name)

    columns = sorted(columns_by_id.values(), key=lambda c: c['_sort_key'])
    for col in columns:
        col.pop('_sort_key', None)

    distinct_dates = sorted({exam_date for (exam_date, _class_id) in cells.keys()})
    rows = []
    for exam_date in distinct_dates:
        row_cells = {}
        for col in columns:
            names = cells.get((exam_date, col['class_id']))
            row_cells[col['class_id']] = ' / '.join(names) if names else ''
        rows.append({
            'date': exam_date.isoformat(),
            'day_name': exam_date.strftime('%A'),
            'cells': row_cells,
        })

    unscheduled.sort(key=lambda item: (item['class_name'], item['subject_name']))

    return {
        'columns': columns,
        'rows': rows,
        'unscheduled': unscheduled,
    }


class ExamGroupViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    required_module = 'examinations'
    queryset = ExamGroup.objects.all()
    permission_classes = [IsAuthenticated, IsSchoolAdmin, HasSchoolAccess]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ExamGroupCreateSerializer
        return ExamGroupSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['school_id'] = _resolve_school_id(self.request)
        return ctx

    def get_queryset(self):
        from django.db.models import Prefetch
        qs = super().get_queryset().select_related(
            'school', 'academic_year', 'term', 'exam_type',
        )

        # Support active/inactive/all filter for exam groups
        exam_is_active = self.request.query_params.get('is_active')
        exam_filter = None
        if exam_is_active is not None:
            if exam_is_active.lower() == 'true':
                exam_filter = True
            elif exam_is_active.lower() == 'false':
                exam_filter = False

        exams_prefetch_qs = Exam.objects.select_related(
            'class_obj', 'exam_type', 'academic_year', 'term',
        ).annotate(
            subjects_count=Count('exam_subjects', filter=Q(exam_subjects__is_active=True)),
        )
        if exam_filter is not None:
            exams_prefetch_qs = exams_prefetch_qs.filter(is_active=exam_filter)

        qs = qs.prefetch_related(
            Prefetch('exams', queryset=exams_prefetch_qs, to_attr='_prefetched_active_exams'),
        )

        academic_year = self.request.query_params.get('academic_year')
        if academic_year:
            qs = qs.filter(academic_year_id=academic_year)
        term = self.request.query_params.get('term')
        if term:
            qs = qs.filter(term_id=term)
        return qs

    def perform_destroy(self, instance):
        instance.delete()  # Cascades to Exam → ExamSubject → StudentMark

    @action(detail=False, methods=['post'], url_path='wizard-create')
    def wizard_create(self, request):
        """Create ExamGroup + per-class Exams + ExamSubjects in one transaction."""
        serializer = ExamGroupWizardCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        school_id = _resolve_school_id(request)
        if not school_id:
            return Response({'detail': 'No school context.'}, status=status.HTTP_400_BAD_REQUEST)

        from django.db import transaction
        from academics.models import ClassSubject
        from students.models import Class

        class_ids = data['class_ids']
        valid_classes = list(Class.objects.filter(
            school_id=school_id, id__in=class_ids, is_active=True,
        ))
        if len(valid_classes) != len(class_ids):
            return Response(
                {'detail': 'One or more class IDs are invalid.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check for conflicts (only active exams block new creation)
        conflicts = []
        for cls in valid_classes:
            existing = Exam.objects.filter(
                school_id=school_id,
                exam_type_id=data['exam_type'],
                class_obj=cls,
                term_id=data.get('term'),
                is_active=True,
            ).first()
            if existing:
                conflicts.append({
                    'class_id': cls.id,
                    'class_name': cls.name,
                    'existing_exam': existing.name,
                })
        if conflicts:
            return Response({
                'detail': 'Some classes already have an active exam of this type for this term.',
                'conflicts': conflicts,
            }, status=status.HTTP_409_CONFLICT)

        # Build lookup: (class_id, subject_id) -> {exam_date, start_time, end_time}
        date_sheet_list = data.get('date_sheet', [])
        date_sheet_map = {}
        for entry in date_sheet_list:
            class_id = entry.get('class_id')
            subject_id = entry.get('subject_id')
            if class_id and subject_id:
                date_sheet_map[(int(class_id), int(subject_id))] = {
                    'exam_date': entry.get('exam_date'),
                    'start_time': entry.get('start_time'),
                    'end_time': entry.get('end_time'),
                }

        default_total = data.get('default_total_marks', 100)
        default_passing = data.get('default_passing_marks', 33)

        with transaction.atomic():
            group = ExamGroup.objects.create(
                school_id=school_id,
                academic_year_id=data['academic_year'],
                term_id=data.get('term'),
                exam_type_id=data['exam_type'],
                name=data['name'],
                description=data.get('description', ''),
                start_date=data.get('start_date'),
                end_date=data.get('end_date'),
            )

            created_exams = []
            for cls in valid_classes:
                exam = Exam.objects.create(
                    school_id=school_id,
                    academic_year_id=data['academic_year'],
                    term_id=data.get('term'),
                    exam_type_id=data['exam_type'],
                    class_obj=cls,
                    exam_group=group,
                    name=f"{data['name']} - {cls.name}",
                    start_date=data.get('start_date'),
                    end_date=data.get('end_date'),
                    status=Exam.Status.SCHEDULED,
                )
                created_exams.append(exam)

            # Per-class subject restriction. A class present here is filtered to
            # exactly its listed subject_ids -- including an empty list, which
            # deliberately yields zero ExamSubjects for that class rather than
            # falling back. A class absent entirely (older clients) keeps every
            # ClassSubject assigned to it.
            subject_restriction_by_class = {
                entry['class_id']: entry.get('subject_ids') or []
                for entry in (data.get('class_subjects') or [])
            }

            all_exam_subjects = []
            for exam in created_exams:
                class_subjects = ClassSubject.objects.filter(
                    school_id=school_id,
                    class_obj=exam.class_obj,
                    is_active=True,
                ).select_related('subject')
                if exam.class_obj_id in subject_restriction_by_class:
                    class_subjects = class_subjects.filter(
                        subject_id__in=subject_restriction_by_class[exam.class_obj_id]
                    )
                for cs in class_subjects:
                    slot = date_sheet_map.get((exam.class_obj_id, cs.subject_id), {})
                    all_exam_subjects.append(ExamSubject(
                        school_id=school_id,
                        exam=exam,
                        subject=cs.subject,
                        total_marks=default_total,
                        passing_marks=default_passing,
                        exam_date=slot.get('exam_date'),
                        start_time=slot.get('start_time'),
                        end_time=slot.get('end_time'),
                    ))

            if all_exam_subjects:
                ExamSubject.objects.bulk_create(all_exam_subjects, ignore_conflicts=True)

        return Response({
            'group_id': group.id,
            'group_name': group.name,
            'exams_created': len(created_exams),
            'subjects_created': len(all_exam_subjects),
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get', 'patch'], url_path='date-sheet')
    def date_sheet(self, request, pk=None):
        """GET: subjects with dates. PATCH: bulk-update exam_date."""
        group = self.get_object()
        school_id = _resolve_school_id(request)

        if request.method == 'GET':
            exam_subjects = ExamSubject.objects.filter(
                exam__exam_group=group,
                exam__is_active=True,
                is_active=True,
                school_id=school_id,
            ).select_related('subject', 'exam', 'exam__class_obj').order_by(
                'subject__name', 'exam__class_obj__grade_level',
            )

            by_subject = {}
            for es in exam_subjects:
                sid = es.subject_id
                if sid not in by_subject:
                    by_subject[sid] = {
                        'subject_id': sid,
                        'subject_name': es.subject.name,
                        'subject_code': es.subject.code,
                        'exam_date': str(es.exam_date) if es.exam_date else None,
                        'classes': [],
                    }
                by_subject[sid]['classes'].append({
                    'exam_subject_id': es.id,
                    'exam_id': es.exam_id,
                    'class_name': es.exam.class_obj.name,
                    'exam_date': str(es.exam_date) if es.exam_date else None,
                    'start_time': str(es.start_time) if es.start_time else None,
                    'end_time': str(es.end_time) if es.end_time else None,
                })

            return Response({
                'group_id': group.id,
                'group_name': group.name,
                'start_date': group.start_date,
                'end_date': group.end_date,
                'subjects': list(by_subject.values()),
            })

        # PATCH
        serializer = DateSheetUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        updated_count = 0
        for entry in serializer.validated_data['date_sheet']:
            es_id = entry.get('exam_subject_id')
            if not es_id:
                continue
            update_fields = {}
            if 'exam_date' in entry:
                update_fields['exam_date'] = entry['exam_date']
            if 'start_time' in entry:
                update_fields['start_time'] = entry['start_time']
            if 'end_time' in entry:
                update_fields['end_time'] = entry['end_time']
            if update_fields:
                count = ExamSubject.objects.filter(
                    id=es_id, exam__exam_group=group, school_id=school_id,
                ).update(**update_fields)
                updated_count += count

        return Response({'updated_count': updated_count})

    @action(detail=True, methods=['post'], url_path='update-date-by-subject')
    def update_date_by_subject(self, request, pk=None):
        """Set the same exam_date for a subject across ALL classes in the group."""
        group = self.get_object()
        school_id = _resolve_school_id(request)
        subject_id = request.data.get('subject_id')
        exam_date = request.data.get('exam_date')

        if not subject_id:
            return Response({'detail': 'subject_id required.'}, status=status.HTTP_400_BAD_REQUEST)

        count = ExamSubject.objects.filter(
            exam__exam_group=group, subject_id=subject_id, school_id=school_id,
        ).update(exam_date=exam_date or None)

        return Response({'updated_count': count})

    @action(detail=True, methods=['get'], url_path='download-date-sheet')
    def download_date_sheet(self, request, pk=None):
        """Generate and return the date sheet as an Excel calendar grid
        (one row per exam date, one column per class)."""
        group = self.get_object()
        school_id = _resolve_school_id(request)
        grid = _build_date_sheet_grid(group, school_id)

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Date Sheet'

        last_col = 2 + len(grid['columns'])  # Date + Day + one per class

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws['A1'] = f'Date Sheet - {group.name}'
        ws['A1'].font = Font(bold=True, size=14)

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        period = ''
        if group.start_date and group.end_date:
            period = f' | {group.start_date} to {group.end_date}'
        ws['A2'] = f'Exam Type: {group.exam_type.name}{period}'
        ws['A2'].font = Font(size=10, color='555555')

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        header_row = 4
        headers = ['Date', 'Day'] + [col['label'] for col in grid['columns']]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        row_idx = header_row + 1
        for row in grid['rows']:
            ws.cell(row=row_idx, column=1, value=row['date']).border = thin_border
            ws.cell(row=row_idx, column=2, value=row['day_name']).border = thin_border
            for col_idx, col in enumerate(grid['columns'], 3):
                value = row['cells'].get(col['class_id'], '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value or '-')
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row_idx += 1

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        for col_idx in range(3, last_col + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

        if grid['unscheduled']:
            row_idx += 1
            ws.cell(row=row_idx, column=1, value='Not yet scheduled:').font = Font(bold=True, size=10)
            for item in grid['unscheduled']:
                row_idx += 1
                ws.cell(row=row_idx, column=1, value=f"{item['subject_name']} ({item['class_name']})")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'DateSheet_{group.name.replace(" ", "_")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='download-date-sheet-pdf')
    def download_date_sheet_pdf(self, request, pk=None):
        """Generate and return the date sheet as a printable PDF calendar grid."""
        from .pdf_generator import DateSheetPDFGenerator

        group = self.get_object()
        school_id = _resolve_school_id(request)
        grid = _build_date_sheet_grid(group, school_id)

        try:
            generator = DateSheetPDFGenerator(group, grid)
            pdf_bytes = generator.generate()

            filename = f'DateSheet_{group.name.replace(" ", "_")}.pdf'
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response(
                {'detail': f'Error generating PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=['post'], url_path='publish-all')
    def publish_all(self, request, pk=None):
        """Publish all exams in the group, notifying the same recipients a single-exam publish would."""
        from notifications.triggers import trigger_exam_result_published

        group = self.get_object()
        exams = list(
            group.exams.filter(is_active=True).select_related('school', 'class_obj', 'academic_year')
        )
        Exam.objects.filter(id__in=[exam.id for exam in exams]).update(status=Exam.Status.PUBLISHED)
        for exam in exams:
            exam.status = Exam.Status.PUBLISHED
            try:
                trigger_exam_result_published(exam)
            except Exception:
                # Do not block publish if notification fanout fails.
                pass
        return Response({'published_count': len(exams)})


class StudentResponseViewSet(ModuleAccessMixin, viewsets.ModelViewSet):
    required_module = 'examinations'
    queryset = StudentResponse.objects.all()
    permission_classes = [IsAuthenticated, HasSchoolAccess]

    def get_serializer_class(self):
        if self.action == 'create':
            return StudentResponseBulkSubmitSerializer
        return StudentResponseSerializer

    def get_queryset(self):
        queryset = self.queryset.select_related('student', 'question', 'exam_paper', 'exam_paper__school')
        school_id = _resolve_school_id(self.request)
        if school_id:
            queryset = queryset.filter(exam_paper__school_id=school_id)
        elif self.request.headers.get('X-School-ID'):
            return queryset.none()

        exam_paper_id = self.request.query_params.get('exam_paper')
        if exam_paper_id:
            queryset = queryset.filter(exam_paper_id=exam_paper_id)

        student_id = self.request.query_params.get('student')
        if student_id:
            queryset = queryset.filter(student_id=student_id)

        question_id = self.request.query_params.get('question')
        if question_id:
            queryset = queryset.filter(question_id=question_id)

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        exam_paper = serializer.validated_data['exam_paper_obj']
        student = serializer.validated_data['student_obj']
        responses = serializer.validated_data['responses']
        school_id = _resolve_school_id(request)
        if school_id and exam_paper.school_id != school_id:
            raise ValidationError({'exam_paper': 'Exam paper is outside the active school context.'})

        created_count = 0
        updated_count = 0
        changed_question_ids = set()
        saved_responses = []

        with transaction.atomic():
            for entry in responses:
                response_obj, created = StudentResponse.objects.update_or_create(
                    student=student,
                    question_id=entry['question'],
                    exam_paper=exam_paper,
                    defaults={
                        'response_text': entry.get('response_text', ''),
                        'marks_awarded': entry.get('marks_awarded'),
                        'is_correct': entry.get('is_correct'),
                        'time_taken_seconds': entry.get('time_taken_seconds'),
                    },
                )
                saved_responses.append(response_obj)
                changed_question_ids.add(response_obj.question_id)
                if created:
                    created_count += 1
                else:
                    updated_count += 1

        from core.task_utils import call_task
        for question_id in changed_question_ids:
            call_task(recompute_question_stats, question_id)

        response_serializer = StudentResponseSerializer(saved_responses, many=True)
        return Response(
            {
                'created_count': created_count,
                'updated_count': updated_count,
                'responses': response_serializer.data,
            },
            status=status.HTTP_201_CREATED,
        )


class ExamViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    required_module = 'examinations'
    queryset = Exam.objects.all()
    permission_classes = [IsAuthenticated, IsSchoolAdminOrReadOnly, HasSchoolAccess]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ExamCreateSerializer
        return ExamSerializer

    def get_permissions(self):
        if self.action in ('bulk_test_preview', 'bulk_test_apply'):
            return [IsAuthenticated(), HasSchoolAccess()]
        return super().get_permissions()

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['school_id'] = _resolve_school_id(self.request)
        return ctx

    def get_queryset(self):
        from academic_sessions.utils import annotate_session_class_display

        qs = super().get_queryset().select_related(
            'school', 'academic_year', 'term', 'exam_type', 'class_obj', 'exam_group',
        ).annotate(
            subjects_count=Count('exam_subjects', filter=Q(exam_subjects__is_active=True)),
        )
        qs = annotate_session_class_display(qs)
        qs = _apply_teacher_exam_scope(qs, self.request, class_field='class_obj_id')
        scope = resolve_class_scope(
            self.request,
            school_id=_resolve_school_id(self.request),
            class_param_names=('class_obj', 'class_id'),
        )
        if scope['invalid']:
            return qs.none()

        academic_year = scope['academic_year_id'] or self.request.query_params.get('academic_year')
        if academic_year:
            qs = qs.filter(academic_year_id=academic_year)
        term = self.request.query_params.get('term')
        if term:
            qs = qs.filter(term_id=term)
        class_obj = scope['class_obj_id']
        if class_obj:
            qs = qs.filter(class_obj_id=class_obj)
        exam_type = self.request.query_params.get('exam_type')
        if exam_type:
            qs = qs.filter(exam_type_id=exam_type)
        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        exam_group = self.request.query_params.get('exam_group')
        if exam_group:
            qs = qs.filter(exam_group_id=exam_group)
        ungrouped = self.request.query_params.get('ungrouped')
        if ungrouped and ungrouped.lower() == 'true':
            qs = qs.filter(exam_group__isnull=True)
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        else:
            qs = qs.filter(is_active=True)
        return qs

    def perform_create(self, serializer):
        super().perform_create(serializer)
        exam = serializer.instance
        # Auto-create ExamSubject entries from the class's assigned subjects
        from academics.models import ClassSubject
        class_subjects = ClassSubject.objects.filter(
            school_id=exam.school_id,
            class_obj=exam.class_obj,
            is_active=True,
        ).select_related('subject')
        exam_subjects = [
            ExamSubject(
                school_id=exam.school_id,
                exam=exam,
                subject=cs.subject,
            )
            for cs in class_subjects
        ]
        if exam_subjects:
            ExamSubject.objects.bulk_create(exam_subjects, ignore_conflicts=True)

    @action(detail=False, methods=['post'], url_path='bulk-test-preview')
    def bulk_test_preview(self, request):
        serializer = BulkTestRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        preview = _build_bulk_test_plan(request, serializer.validated_data)
        return Response(preview)

    @action(detail=False, methods=['post'], url_path='bulk-test-apply')
    def bulk_test_apply(self, request):
        serializer = BulkTestRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        preview = _build_bulk_test_plan(request, serializer.validated_data)
        if not preview['can_apply']:
            return Response(
                {
                    **preview,
                    'detail': 'Preview contains conflicts or inaccessible subjects. Resolve them before applying.',
                },
                status=status.HTTP_409_CONFLICT,
            )

        school_id = _resolve_school_id(request)
        created_tests = []
        try:
            with transaction.atomic():
                for item in preview['tests']:
                    exam = Exam.objects.create(
                        school_id=school_id,
                        academic_year_id=preview['academic_year'],
                        term_id=preview['term'],
                        exam_type_id=preview['exam_type'],
                        class_obj_id=preview['class_obj'],
                        exam_group=None,
                        name=item['name'],
                        start_date=item['exam_date'],
                        end_date=item['exam_date'],
                        status=Exam.Status.SCHEDULED,
                    )
                    ExamSubject.objects.create(
                        school_id=school_id,
                        exam=exam,
                        subject_id=item['subject_id'],
                        total_marks=item.get('total_marks') or Decimal('100.00'),
                        passing_marks=((item.get('total_marks') or Decimal('100.00')) * Decimal('0.33')).quantize(Decimal('0.01')),
                        exam_date=item['exam_date'],
                        start_time=item.get('start_time'),
                        end_time=item.get('end_time'),
                    )
                    created_tests.append({
                        'exam_id': exam.id,
                        'name': exam.name,
                        'subject_id': item['subject_id'],
                        'subject_name': item['subject_name'],
                        'exam_date': item['exam_date'],
                        'total_marks': item.get('total_marks') or Decimal('100.00'),
                    })
        except IntegrityError as exc:
            message = str(exc)
            legacy_constraint = 'examinations_exam_school_id_exam_type_id_c_bf67c535_uniq'
            if legacy_constraint in message:
                return Response(
                    {
                        **preview,
                        'detail': (
                            'Your database still enforces the legacy standalone-test uniqueness constraint. '
                            'Run examinations migrations (including 0015+) and retry.'
                        ),
                        'constraint': legacy_constraint,
                    },
                    status=status.HTTP_409_CONFLICT,
                )
            return Response(
                {
                    **preview,
                    'detail': 'A database integrity error occurred while creating tests. Please retry.',
                },
                status=status.HTTP_409_CONFLICT,
            )

        return Response({
            'created_count': len(created_tests),
            'created_tests': created_tests,
            'class_name': preview['class_name'],
            'exam_type_name': preview['exam_type_name'],
            'academic_year_name': preview['academic_year_name'],
            'term_name': preview['term_name'],
        }, status=status.HTTP_201_CREATED)

    def perform_destroy(self, instance):
        instance.delete()  # Cascades to ExamSubject → StudentMark

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        exam = self.get_object()
        exam.status = Exam.Status.PUBLISHED
        exam.save(update_fields=['status'])
        try:
            from notifications.triggers import trigger_exam_result_published
            trigger_exam_result_published(exam)
        except Exception:
            # Do not block publish if notification fanout fails.
            pass
        return Response(ExamSerializer(exam).data)

    @action(detail=True, methods=['post'], url_path='generate-comments')
    def generate_comments(self, request, pk=None):
        """AI: Generate personalized report card comments for all marks in this exam.

        Uses AI to generate 2-3 sentence comments based on each student's marks,
        grade, and attendance record. Comments can be edited by teachers after generation.
        Skips marks that already have AI comments (use force=true to regenerate all).
        """
        exam = self.get_object()
        school_id = _resolve_school_id(request)
        force = request.data.get('force', False)

        if not school_id:
            return Response({'detail': 'No school selected.'}, status=status.HTTP_400_BAD_REQUEST)

        # If force=true, clear existing AI comments first
        if force:
            StudentMark.objects.filter(
                exam_subject__exam=exam,
                school_id=school_id,
            ).update(ai_comment='', ai_comment_generated_at=None)

        from schools.models import School
        school = School.objects.get(id=school_id)

        from .ai_comments_service import ReportCardCommentGenerator
        generator = ReportCardCommentGenerator(school)
        result = generator.generate_for_exam(exam.id)

        return Response(result)

    @action(detail=True, methods=['post'], url_path='populate-subjects')
    def populate_subjects(self, request, pk=None):
        """Re-sync exam subjects from the class's current ClassSubject assignments."""
        exam = self.get_object()
        school_id = _resolve_school_id(request)

        from academics.models import ClassSubject
        class_subjects = ClassSubject.objects.filter(
            school_id=school_id,
            class_obj=exam.class_obj,
            is_active=True,
        ).select_related('subject')

        existing_subject_ids = set(
            exam.exam_subjects.filter(is_active=True).values_list('subject_id', flat=True)
        )

        new_exam_subjects = [
            ExamSubject(school_id=school_id, exam=exam, subject=cs.subject)
            for cs in class_subjects
            if cs.subject_id not in existing_subject_ids
        ]

        created = []
        if new_exam_subjects:
            created = ExamSubject.objects.bulk_create(new_exam_subjects, ignore_conflicts=True)

        return Response({
            'added_count': len(created),
            'total_count': exam.exam_subjects.filter(is_active=True).count(),
        })

    @action(detail=True, methods=['get'])
    def results(self, request, pk=None):
        exam = self.get_object()
        school_id = _resolve_school_id(request)
        exam_subjects = exam.exam_subjects.filter(is_active=True).select_related('subject')

        from students.models import Student
        students = Student.objects.filter(
            school_id=school_id,
            class_obj=exam.class_obj,
            is_active=True,
        ).order_by('roll_number')

        grade_scales = list(GradeScale.objects.filter(
            school_id=school_id, is_active=True,
        ).order_by('-min_percentage'))

        # Prefetch all marks in one query and build lookup dict
        all_marks = StudentMark.objects.filter(
            exam_subject__in=exam_subjects, school_id=school_id,
        ).select_related('exam_subject')
        marks_lookup = {
            (m.student_id, m.exam_subject_id): m for m in all_marks
        }

        results = []
        for student in students:
            marks_list = []
            total_obtained = Decimal('0')
            total_possible = Decimal('0')
            all_pass = True

            for es in exam_subjects:
                mark = marks_lookup.get((student.id, es.id))
                obtained = mark.marks_obtained if mark and not mark.is_absent else None
                is_absent = mark.is_absent if mark else False

                marks_list.append({
                    'subject_id': es.subject_id,
                    'subject_name': es.subject.name,
                    'total_marks': float(es.total_marks),
                    'passing_marks': float(es.passing_marks),
                    'marks_obtained': float(obtained) if obtained is not None else None,
                    'is_absent': is_absent,
                    'is_pass': obtained is not None and obtained >= es.passing_marks,
                    'ai_comment': mark.ai_comment if mark else '',
                })

                if obtained is not None:
                    total_obtained += obtained
                    total_possible += es.total_marks
                    if obtained < es.passing_marks:
                        all_pass = False
                else:
                    total_possible += es.total_marks
                    all_pass = False

            percentage = float(total_obtained / total_possible * 100) if total_possible > 0 else 0
            grade_label = self._get_grade(percentage, grade_scales)

            results.append({
                'student_id': student.id,
                'student_name': student.name,
                'roll_number': student.roll_number,
                'marks': marks_list,
                'total_obtained': float(total_obtained),
                'total_possible': float(total_possible),
                'percentage': round(percentage, 2),
                'grade': grade_label,
                'is_pass': all_pass,
            })

        # Calculate ranks
        results.sort(key=lambda x: x['percentage'], reverse=True)
        for i, r in enumerate(results):
            r['rank'] = i + 1

        return Response({
            'exam': ExamSerializer(exam).data,
            'exam_type_weight': float(exam.exam_type.weight),
            'subjects': ExamSubjectSerializer(exam_subjects, many=True).data,
            'results': results,
        })

    @action(detail=True, methods=['get'])
    def class_summary(self, request, pk=None):
        exam = self.get_object()
        school_id = _resolve_school_id(request)
        exam_subjects = exam.exam_subjects.filter(is_active=True).select_related('subject')

        from students.models import Student
        students = Student.objects.filter(
            school_id=school_id,
            class_obj=exam.class_obj,
            is_active=True,
        )

        # Prefetch all marks for this exam in one query
        all_marks = StudentMark.objects.filter(
            exam_subject__in=exam_subjects, school_id=school_id,
            is_absent=False, marks_obtained__isnull=False,
        )
        # Group marks by exam_subject_id
        marks_by_subject = {}
        for m in all_marks:
            marks_by_subject.setdefault(m.exam_subject_id, []).append(m)

        subject_stats = []
        for es in exam_subjects:
            subject_marks = marks_by_subject.get(es.id, [])
            marks_values = [float(m.marks_obtained) for m in subject_marks]
            passed = sum(1 for m in subject_marks if m.marks_obtained >= es.passing_marks)
            subject_stats.append({
                'subject_name': es.subject.name,
                'total_marks': float(es.total_marks),
                'students_appeared': len(marks_values),
                'average': round(sum(marks_values) / len(marks_values), 2) if marks_values else 0,
                'highest': max(marks_values) if marks_values else 0,
                'lowest': min(marks_values) if marks_values else 0,
                'passed': passed,
                'failed': len(marks_values) - passed,
            })

        return Response({
            'exam': ExamSerializer(exam).data,
            'total_students': students.count(),
            'subject_stats': subject_stats,
        })

    def _get_grade(self, percentage, grade_scales):
        for gs in grade_scales:
            if float(gs.min_percentage) <= percentage <= float(gs.max_percentage):
                return gs.grade_label
        return '-'


class ExamSubjectViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    required_module = 'examinations'
    queryset = ExamSubject.objects.all()
    permission_classes = [IsAuthenticated, IsSchoolAdminOrReadOnly, HasSchoolAccess]


    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ExamSubjectCreateSerializer
        return ExamSubjectSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['school_id'] = _resolve_school_id(self.request)
        return ctx

    def get_queryset(self):
        qs = super().get_queryset().select_related('school', 'exam', 'subject')
        qs = _apply_teacher_exam_scope(
            qs,
            self.request,
            class_field='exam__class_obj_id',
            subject_field='subject_id',
        )
        scope = resolve_class_scope(
            self.request,
            school_id=_resolve_school_id(self.request),
            class_param_names=('class_obj', 'class_id'),
        )
        if scope['invalid']:
            return qs.none()

        exam = self.request.query_params.get('exam')
        if exam:
            qs = qs.filter(exam_id=exam)
        class_obj = scope['class_obj_id']
        if class_obj:
            qs = qs.filter(exam__class_obj_id=class_obj)
        academic_year = scope['academic_year_id'] or self.request.query_params.get('academic_year')
        if academic_year:
            qs = qs.filter(exam__academic_year_id=academic_year)
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        else:
            qs = qs.filter(is_active=True)
        return qs

    def perform_destroy(self, instance):
        instance.delete()  # Cascades to StudentMark


class StudentMarkViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    required_module = 'examinations'
    queryset = StudentMark.objects.all()
    permission_classes = [IsAuthenticated, IsSchoolAdminOrReadOnly, HasSchoolAccess]


    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return StudentMarkCreateSerializer
        return StudentMarkSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['school_id'] = _resolve_school_id(self.request)
        return ctx

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            'school', 'exam_subject', 'exam_subject__subject',
            'exam_subject__exam', 'student',
        )
        qs = _apply_teacher_exam_scope(
            qs,
            self.request,
            class_field='exam_subject__exam__class_obj_id',
            subject_field='exam_subject__subject_id',
        )
        scope = resolve_class_scope(
            self.request,
            school_id=_resolve_school_id(self.request),
            class_param_names=('class_obj', 'class_id'),
        )
        if scope['invalid']:
            return qs.none()

        exam_subject = self.request.query_params.get('exam_subject')
        if exam_subject:
            qs = qs.filter(exam_subject_id=exam_subject)
        student = self.request.query_params.get('student')
        if student:
            qs = qs.filter(student_id=student)
        class_obj = scope['class_obj_id']
        if class_obj:
            qs = qs.filter(exam_subject__exam__class_obj_id=class_obj)

        academic_year = scope['academic_year_id'] or self.request.query_params.get('academic_year')
        if academic_year:
            qs = qs.filter(exam_subject__exam__academic_year_id=academic_year)
        return qs

    @action(detail=False, methods=['post'])
    def bulk_entry(self, request):
        serializer = StudentMarkBulkEntrySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        school_id = _resolve_school_id(request)
        exam_subject_id = serializer.validated_data['exam_subject_id']
        marks_data = serializer.validated_data['marks']

        try:
            exam_subject = ExamSubject.objects.get(
                pk=exam_subject_id, school_id=school_id,
            )
        except ExamSubject.DoesNotExist:
            return Response(
                {'detail': 'Exam subject not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        created = 0
        updated = 0
        errors = []

        from academic_sessions.models import StudentEnrollment

        for entry in marks_data:
            student_id = entry.get('student_id')
            marks_obtained = entry.get('marks_obtained')
            is_absent = entry.get('is_absent', False)
            remarks = entry.get('remarks', '')

            if marks_obtained is not None:
                marks_obtained = Decimal(str(marks_obtained))

            enrollment = StudentEnrollment.objects.filter(
                school_id=school_id,
                student_id=student_id,
                academic_year_id=exam_subject.exam.academic_year_id,
                class_obj_id=exam_subject.exam.class_obj_id,
            ).order_by('-is_active', '-created_at').first()

            try:
                mark, was_created = StudentMark.objects.update_or_create(
                    school_id=school_id,
                    exam_subject=exam_subject,
                    student_id=student_id,
                    defaults={
                        'marks_obtained': None if is_absent else marks_obtained,
                        'is_absent': is_absent,
                        'remarks': remarks,
                        'enrollment': enrollment,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({'student_id': student_id, 'error': str(e)})

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
            'message': f'{created + updated} marks saved.',
        })

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Generate Excel template pre-filled with student names for marks entry."""
        school_id = _resolve_school_id(request)
        exam_subject_id = request.query_params.get('exam_subject_id')
        if not exam_subject_id:
            return Response(
                {'detail': 'exam_subject_id param required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            exam_subject = ExamSubject.objects.select_related(
                'exam', 'exam__class_obj', 'subject',
            ).get(pk=exam_subject_id, school_id=school_id)
        except ExamSubject.DoesNotExist:
            return Response(
                {'detail': 'Exam subject not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Get students enrolled in the exam's class for the exam's academic year
        from students.models import Student
        from academic_sessions.models import StudentEnrollment
        students = Student.objects.filter(
            school_id=school_id,
            class_obj=exam_subject.exam.class_obj,
            is_active=True,
        )
        # Filter by enrollment if the school uses enrollments
        academic_year_id = exam_subject.exam.academic_year_id
        if academic_year_id and StudentEnrollment.objects.filter(school_id=school_id).exists():
            enrolled_ids = StudentEnrollment.objects.filter(
                academic_year_id=academic_year_id,
                is_active=True,
            ).values_list('student_id', flat=True)
            students = students.filter(id__in=enrolled_ids)
        students = students.order_by('roll_number', 'name')

        # Also check for existing marks
        existing_marks = {
            m.student_id: m
            for m in StudentMark.objects.filter(
                school_id=school_id,
                exam_subject=exam_subject,
            )
        }

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Marks Entry'

        # Header info rows
        header_font = Font(bold=True, size=12)
        info_font = Font(size=10, color='555555')
        ws.merge_cells('A1:E1')
        ws['A1'] = f'Marks Entry - {exam_subject.exam.name}'
        ws['A1'].font = header_font
        from academic_sessions.utils import resolve_class_display_name

        ws.merge_cells('A2:E2')
        ws['A2'] = (
            f'Subject: {exam_subject.subject.name} | '
            f'Class: {resolve_class_display_name(exam_subject.exam.school_id, exam_subject.exam.academic_year_id, exam_subject.exam.class_obj)} | '
            f'Total Marks: {exam_subject.total_marks} | '
            f'Passing: {exam_subject.passing_marks}'
        )
        ws['A2'].font = info_font

        # Hidden metadata row for upload parsing
        ws['A3'] = 'exam_subject_id'
        ws['B3'] = str(exam_subject.id)
        ws.row_dimensions[3].hidden = True

        # Column headers
        headers = ['Student ID', 'Roll Number', 'Student Name', f'Marks (out of {exam_subject.total_marks})', 'Absent (Y/N)', 'Remarks']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Student rows
        for row_idx, student in enumerate(students, 5):
            existing = existing_marks.get(student.id)
            ws.cell(row=row_idx, column=1, value=student.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=student.roll_number or '').border = thin_border
            name_cell = ws.cell(row=row_idx, column=3, value=student.name)
            name_cell.border = thin_border
            name_cell.font = Font(size=10)

            marks_cell = ws.cell(row=row_idx, column=4)
            if existing and existing.marks_obtained is not None:
                marks_cell.value = float(existing.marks_obtained)
            marks_cell.border = thin_border
            marks_cell.alignment = Alignment(horizontal='center')

            absent_cell = ws.cell(row=row_idx, column=5)
            if existing and existing.is_absent:
                absent_cell.value = 'Y'
            absent_cell.border = thin_border
            absent_cell.alignment = Alignment(horizontal='center')

            remarks_cell = ws.cell(row=row_idx, column=6)
            if existing and existing.remarks:
                remarks_cell.value = existing.remarks
            remarks_cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 25

        # Lock student ID and name columns (read-only visual cue)
        lock_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for row_idx in range(5, 5 + students.count()):
            ws.cell(row=row_idx, column=1).fill = lock_fill
            ws.cell(row=row_idx, column=2).fill = lock_fill
            ws.cell(row=row_idx, column=3).fill = lock_fill

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = (
            f'Marks_Template_{exam_subject.exam.name}_'
            f'{exam_subject.subject.code}.xlsx'
        ).replace(' ', '_')

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'])
    def by_student(self, request):
        student_id = request.query_params.get('student_id')
        if not student_id:
            return Response(
                {'detail': 'student_id param required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        qs = self.get_queryset().filter(student_id=student_id)
        serializer = StudentMarkSerializer(qs, many=True)
        return Response(serializer.data)


class GradeScaleViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    required_module = 'examinations'
    queryset = GradeScale.objects.all()
    permission_classes = [IsAuthenticated, IsSchoolAdmin, HasSchoolAccess]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return GradeScaleCreateSerializer
        return GradeScaleSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['school_id'] = _resolve_school_id(self.request)
        return ctx

    def get_queryset(self):
        qs = super().get_queryset()
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        else:
            qs = qs.filter(is_active=True)
        return qs

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


class StudentTermAssessmentView(ModuleAccessMixin, APIView):
    """
    Skills/behaviour ratings + remarks for one student's academic year + month (upsert).
    GET  ?student_id=&academic_year=&month= -> existing row, or a blank shape if none yet.
    POST {student, academic_year, month, ...ratings, teacher_remark, principal_remark} -> create/update.
    """
    required_module = 'examinations'
    permission_classes = [IsAuthenticated, CanManageStudentAssessments, HasSchoolAccess]

    RATING_FIELDS = [
        'listening', 'speaking', 'writing', 'reading', 'participation', 'confidence', 'social_skills',
        'discipline', 'respect', 'teamwork', 'class_participation', 'responsibility',
    ]

    @staticmethod
    def _parse_month(raw_month):
        try:
            month = int(raw_month)
        except (TypeError, ValueError):
            return None
        return month if 1 <= month <= 12 else None

    def _teacher_can_access_class(self, request, school_id, academic_year_id, class_obj_id=None, session_class_id=None):
        role = get_effective_role(request)
        if role in ADMIN_ROLES:
            return True
        if role != 'TEACHER':
            return False

        scope = get_teacher_combined_scope(request, school_id=school_id, academic_year_id=academic_year_id)
        allowed_class_ids = set(scope.get('all_class_ids', set()))
        allowed_session_ids = set(scope.get('full_session_class_ids', set()))

        if session_class_id:
            return int(session_class_id) in allowed_session_ids or int(class_obj_id or 0) in allowed_class_ids
        if class_obj_id:
            return int(class_obj_id) in allowed_class_ids
        return False

    def _teacher_can_access_student(self, request, school_id, student_id, academic_year_id):
        role = get_effective_role(request)
        if role in ADMIN_ROLES:
            return True
        if role != 'TEACHER':
            return False

        from academic_sessions.models import StudentEnrollment

        enrollment = StudentEnrollment.objects.filter(
            school_id=school_id,
            student_id=student_id,
            academic_year_id=academic_year_id,
            is_active=True,
        ).select_related('session_class', 'class_obj').first()
        if not enrollment:
            return False

        return self._teacher_can_access_class(
            request,
            school_id,
            academic_year_id,
            class_obj_id=enrollment.class_obj_id,
            session_class_id=enrollment.session_class_id,
        )

    def _build_defaults(self, request, row, existing=None):
        role = get_effective_role(request)
        defaults = {
            field: (None if row.get(field) in (None, '') else row.get(field))
            for field in self.RATING_FIELDS
        }
        defaults['teacher_remark'] = row.get('teacher_remark', '')
        if role in ADMIN_ROLES:
            defaults['principal_remark'] = row.get('principal_remark', '')
        elif existing is not None:
            defaults['principal_remark'] = existing.principal_remark or ''
        else:
            defaults['principal_remark'] = ''
        defaults['updated_by'] = request.user
        return defaults

    def _reject_no_access(self):
        return Response({'detail': 'You are not assigned to this student/class for the selected academic year.'}, status=403)

    def get(self, request):
        school_id = _resolve_school_id(request)
        if not school_id:
            return Response({'error': 'school_id required'}, status=400)

        student_id = request.query_params.get('student_id')
        academic_year_id = request.query_params.get('academic_year')
        month = self._parse_month(request.query_params.get('month'))
        if not student_id or not academic_year_id or month is None:
            return Response({'error': 'student_id, academic_year, and month (1-12) are required'}, status=400)

        if not self._teacher_can_access_student(request, school_id, student_id, academic_year_id):
            return self._reject_no_access()

        assessment = StudentTermAssessment.objects.filter(
            school_id=school_id,
            student_id=student_id,
            academic_year_id=academic_year_id,
            month=month,
        ).first()
        if not assessment:
            blank = {f: None for f in self.RATING_FIELDS}
            blank.update({
                'exists': False,
                'student': int(student_id),
                'academic_year': int(academic_year_id),
                'month': month,
                'teacher_remark': '', 'principal_remark': '',
            })
            return Response(blank)

        data = StudentTermAssessmentSerializer(assessment).data
        data['exists'] = True
        return Response(data)

    def post(self, request):
        school_id = _resolve_school_id(request)
        if not school_id:
            return Response({'error': 'school_id required'}, status=400)

        student_id = request.data.get('student')
        academic_year_id = request.data.get('academic_year')
        month = self._parse_month(request.data.get('month'))
        if not student_id or not academic_year_id or month is None:
            return Response({'error': 'student, academic_year, and month (1-12) are required'}, status=400)

        if not self._teacher_can_access_student(request, school_id, student_id, academic_year_id):
            return self._reject_no_access()

        existing = StudentTermAssessment.objects.filter(
            school_id=school_id,
            student_id=student_id,
            academic_year_id=academic_year_id,
            month=month,
        ).first()
        defaults = self._build_defaults(request, request.data, existing=existing)

        assessment, _created = StudentTermAssessment.objects.update_or_create(
            school_id=school_id,
            student_id=student_id,
            academic_year_id=academic_year_id,
            month=month,
            defaults=defaults,
        )
        data = StudentTermAssessmentSerializer(assessment).data
        data['exists'] = True
        return Response(data)


class StudentTermAssessmentRosterView(ModuleAccessMixin, APIView):
    """
    Class/month roster for student assessments.
    GET ?academic_year=&month=&session_class= (or class_obj=) -> one row per enrolled student
    with existing monthly assessment data merged in.
    """
    required_module = 'examinations'
    permission_classes = [IsAuthenticated, CanManageStudentAssessments, HasSchoolAccess]

    RATING_FIELDS = StudentTermAssessmentView.RATING_FIELDS

    def get(self, request):
        school_id = _resolve_school_id(request)
        if not school_id:
            return Response({'error': 'school_id required'}, status=400)

        academic_year_id = request.query_params.get('academic_year')
        month = StudentTermAssessmentView._parse_month(request.query_params.get('month'))
        session_class_id = request.query_params.get('session_class')
        class_obj_id = request.query_params.get('class_obj')

        if not academic_year_id or month is None:
            return Response({'error': 'academic_year and month (1-12) are required'}, status=400)
        if not session_class_id and not class_obj_id:
            return Response({'error': 'session_class or class_obj is required'}, status=400)

        if not StudentTermAssessmentView()._teacher_can_access_class(
            request,
            school_id,
            academic_year_id,
            class_obj_id=class_obj_id,
            session_class_id=session_class_id,
        ):
            return Response({'detail': 'You are not assigned to this class for the selected academic year.'}, status=403)

        from academic_sessions.models import StudentEnrollment

        enrollments = StudentEnrollment.objects.filter(
            school_id=school_id,
            academic_year_id=academic_year_id,
            is_active=True,
        ).select_related('student', 'session_class', 'class_obj')

        if session_class_id:
            enrollments = enrollments.filter(session_class_id=session_class_id)
        else:
            enrollments = enrollments.filter(class_obj_id=class_obj_id)

        enrollments = enrollments.order_by('roll_number', 'student__name', 'id')
        student_ids = list(enrollments.values_list('student_id', flat=True))

        assessments = StudentTermAssessment.objects.filter(
            school_id=school_id,
            academic_year_id=academic_year_id,
            month=month,
            student_id__in=student_ids,
        ).select_related('updated_by')
        assessments_by_student = {row.student_id: row for row in assessments}

        results = []
        for enrollment in enrollments:
            assessment = assessments_by_student.get(enrollment.student_id)
            if assessment:
                row = StudentTermAssessmentSerializer(assessment).data
                row['exists'] = True
            else:
                row = {f: None for f in self.RATING_FIELDS}
                row.update({
                    'exists': False,
                    'student': enrollment.student_id,
                    'academic_year': int(academic_year_id),
                    'month': month,
                    'teacher_remark': '',
                    'principal_remark': '',
                    'updated_by': None,
                    'updated_by_name': None,
                    'updated_at': None,
                })

            row['student_name'] = enrollment.student.name
            row['roll_number'] = enrollment.roll_number
            row['enrollment_id'] = enrollment.id
            row['class_obj'] = enrollment.class_obj_id
            row['session_class'] = enrollment.session_class_id
            results.append(row)

        return Response({
            'academic_year': int(academic_year_id),
            'month': month,
            'class_obj': int(class_obj_id) if class_obj_id else None,
            'session_class': int(session_class_id) if session_class_id else None,
            'count': len(results),
            'results': results,
        })


class StudentTermAssessmentBulkSaveView(ModuleAccessMixin, APIView):
    """
    Bulk upsert student assessments for a selected class + month.
    POST {
      academic_year, month, session_class|class_obj,
      assessments: [{student, ...rating_fields, teacher_remark, principal_remark}]
    }
    """
    required_module = 'examinations'
    permission_classes = [IsAuthenticated, CanManageStudentAssessments, HasSchoolAccess]

    RATING_FIELDS = StudentTermAssessmentView.RATING_FIELDS

    def post(self, request):
        school_id = _resolve_school_id(request)
        if not school_id:
            return Response({'error': 'school_id required'}, status=400)

        academic_year_id = request.data.get('academic_year')
        month = StudentTermAssessmentView._parse_month(request.data.get('month'))
        session_class_id = request.data.get('session_class')
        class_obj_id = request.data.get('class_obj')
        rows = request.data.get('assessments') or []

        if not academic_year_id or month is None:
            return Response({'error': 'academic_year and month (1-12) are required'}, status=400)
        if not session_class_id and not class_obj_id:
            return Response({'error': 'session_class or class_obj is required'}, status=400)
        if not isinstance(rows, list):
            return Response({'error': 'assessments must be a list'}, status=400)

        if not StudentTermAssessmentView()._teacher_can_access_class(
            request,
            school_id,
            academic_year_id,
            class_obj_id=class_obj_id,
            session_class_id=session_class_id,
        ):
            return Response({'detail': 'You are not assigned to this class for the selected academic year.'}, status=403)

        from academic_sessions.models import StudentEnrollment

        enrollments = StudentEnrollment.objects.filter(
            school_id=school_id,
            academic_year_id=academic_year_id,
            is_active=True,
        )
        if session_class_id:
            enrollments = enrollments.filter(session_class_id=session_class_id)
        else:
            enrollments = enrollments.filter(class_obj_id=class_obj_id)

        allowed_student_ids = set(enrollments.values_list('student_id', flat=True))
        if not allowed_student_ids:
            return Response({
                'error': 'No active enrollments found for selected class and academic year',
            }, status=400)

        created = 0
        updated = 0
        errors = []
        saved_results = []

        existing_map = {
            row.student_id: row
            for row in StudentTermAssessment.objects.filter(
                school_id=school_id,
                academic_year_id=academic_year_id,
                month=month,
                student_id__in=allowed_student_ids,
            )
        }

        with transaction.atomic():
            for index, row in enumerate(rows):
                student_id = row.get('student')
                if not student_id:
                    errors.append({'index': index, 'error': 'student is required'})
                    continue
                if student_id not in allowed_student_ids:
                    errors.append({
                        'index': index,
                        'student': student_id,
                        'error': 'student is not enrolled in selected class/academic_year',
                    })
                    continue

                defaults = StudentTermAssessmentView()._build_defaults(request, row, existing=existing_map.get(student_id))

                obj, was_created = StudentTermAssessment.objects.update_or_create(
                    school_id=school_id,
                    student_id=student_id,
                    academic_year_id=academic_year_id,
                    month=month,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

                payload = StudentTermAssessmentSerializer(obj).data
                payload['exists'] = True
                saved_results.append(payload)

        return Response({
            'academic_year': int(academic_year_id),
            'month': month,
            'class_obj': int(class_obj_id) if class_obj_id else None,
            'session_class': int(session_class_id) if session_class_id else None,
            'submitted_count': len(rows),
            'created': created,
            'updated': updated,
            'error_count': len(errors),
            'errors': errors,
            'results': saved_results,
        })


class StudentTermAssessmentAIRemarkView(ModuleAccessMixin, APIView):
    """
    Draft a teacher/principal remark from a student's current skill/behaviour ratings.
    POST {ratings: {field_label: rating_value(1-5)}, remark_type: 'teacher'|'principal'} -> {remark, fallback}
    """
    required_module = 'examinations'
    permission_classes = [IsAuthenticated, CanManageStudentAssessments, HasSchoolAccess]

    def post(self, request):
        school_id = _resolve_school_id(request)
        if not school_id:
            return Response({'error': 'school_id required'}, status=400)

        ratings = request.data.get('ratings')
        remark_type = request.data.get('remark_type') or 'teacher'
        if not isinstance(ratings, dict):
            return Response({'error': 'ratings (object) is required'}, status=400)
        if remark_type not in ('teacher', 'principal'):
            return Response({'error': "remark_type must be 'teacher' or 'principal'"}, status=400)

        from .ai_comments_service import generate_term_assessment_remark

        remark, fallback = generate_term_assessment_remark(ratings, remark_type=remark_type)
        if remark is None:
            return Response(
                {'error': 'Rate at least one skill or behaviour before requesting an AI suggestion.'},
                status=400,
            )
        return Response({'remark': remark, 'fallback': fallback})


class ReportCardView(ModuleAccessMixin, APIView):
    required_module = 'examinations'
    permission_classes = [IsAuthenticated, IsSchoolAdminOrReadOnly, HasSchoolAccess]

    def get(self, request):
        student_id = request.query_params.get('student_id')
        academic_year_id = request.query_params.get('academic_year_id')
        term_id = request.query_params.get('term_id')
        enrollment_id = request.query_params.get('enrollment_id')

        if not student_id:
            return Response({'detail': 'student_id required.'}, status=400)
        if not academic_year_id and not enrollment_id:
            return Response(
                {'detail': 'academic_year_id or enrollment_id is required.'},
                status=400,
            )

        school_id = _resolve_school_id(request)

        from students.models import Student
        from academic_sessions.models import StudentEnrollment
        try:
            student = Student.objects.select_related('class_obj', 'school').get(
                pk=student_id, school_id=school_id,
            )
        except Student.DoesNotExist:
            return Response({'detail': 'Student not found.'}, status=404)

        enrollment_qs = StudentEnrollment.objects.select_related('class_obj', 'academic_year', 'session_class').filter(
            school_id=school_id,
            student_id=student.id,
        )
        if enrollment_id:
            enrollment_qs = enrollment_qs.filter(pk=enrollment_id)
        else:
            enrollment_qs = enrollment_qs.filter(academic_year_id=academic_year_id)

        enrollment = enrollment_qs.order_by('-created_at').first()
        if not enrollment:
            return Response(
                {'detail': 'No enrollment found for the selected student/session.'},
                status=404,
            )

        # Get exams for the class captured in the selected enrollment/session.
        exam_filter = {
            'school_id': school_id,
            'class_obj': enrollment.class_obj,
            'is_active': True,
            'status': Exam.Status.PUBLISHED,
            'academic_year_id': enrollment.academic_year_id,
        }
        if term_id:
            exam_filter['term_id'] = term_id

        exams = Exam.objects.filter(**exam_filter).select_related(
            'exam_type', 'academic_year', 'term',
        ).order_by('start_date')

        grade_scales = list(GradeScale.objects.filter(
            school_id=school_id, is_active=True,
        ).order_by('-min_percentage'))

        # Prefetch all exam subjects and marks for this student in one query
        all_exam_subjects = ExamSubject.objects.filter(
            exam__in=exams, is_active=True,
        ).select_related('subject')
        student_marks = StudentMark.objects.filter(
            exam_subject__in=all_exam_subjects,
            student=student,
            school_id=school_id,
        )
        marks_lookup = {m.exam_subject_id: m for m in student_marks}

        # Group exam subjects by exam
        es_by_exam = {}
        for es in all_exam_subjects:
            es_by_exam.setdefault(es.exam_id, []).append(es)

        all_subjects = {}
        exam_data = []

        for exam in exams:
            exam_subjects = es_by_exam.get(exam.id, [])
            exam_marks = {}

            for es in exam_subjects:
                if es.subject_id not in all_subjects:
                    all_subjects[es.subject_id] = es.subject.name

                mark = marks_lookup.get(es.id)
                exam_marks[es.subject_id] = {
                    'total_marks': float(es.total_marks),
                    'marks_obtained': float(mark.marks_obtained) if mark and mark.marks_obtained else None,
                    'is_absent': mark.is_absent if mark else False,
                    'ai_comment': mark.ai_comment if mark else '',
                }

            exam_data.append({
                'exam_id': exam.id,
                'exam_name': exam.name,
                'exam_type': exam.exam_type.name,
                'term': exam.term.name if exam.term else None,
                'marks': exam_marks,
            })

        # Determine weighted vs simple calculation
        from schools.models import School
        school = School.objects.get(pk=school_id)
        use_weighted = (school.exam_config or {}).get('weighted_average_enabled', False)

        # Calculate overall totals
        grand_total_obtained = Decimal('0')
        grand_total_possible = Decimal('0')

        if use_weighted and exams.count() > 1:
            # Weighted: group by exam_type, compute per-type percentage, apply weights
            exam_type_data = {}
            for exam in exams:
                et_id = exam.exam_type_id
                if et_id not in exam_type_data:
                    exam_type_data[et_id] = {
                        'weight': exam.exam_type.weight,
                        'obtained': Decimal('0'),
                        'possible': Decimal('0'),
                    }
                for es_item in es_by_exam.get(exam.id, []):
                    mark = marks_lookup.get(es_item.id)
                    if mark and mark.marks_obtained is not None and not mark.is_absent:
                        exam_type_data[et_id]['obtained'] += mark.marks_obtained
                    exam_type_data[et_id]['possible'] += es_item.total_marks

            total_weight = sum(d['weight'] for d in exam_type_data.values() if d['possible'] > 0)
            if total_weight > 0:
                weighted_sum = Decimal('0')
                for data in exam_type_data.values():
                    if data['possible'] > 0:
                        type_pct = data['obtained'] / data['possible'] * 100
                        weighted_sum += type_pct * (data['weight'] / total_weight)
                overall_pct = float(weighted_sum)
            else:
                overall_pct = 0

            grand_total_obtained = sum((d['obtained'] for d in exam_type_data.values()), Decimal('0'))
            grand_total_possible = sum((d['possible'] for d in exam_type_data.values()), Decimal('0'))
        else:
            # Simple average
            for es_item in all_exam_subjects:
                mark = marks_lookup.get(es_item.id)
                if mark and mark.marks_obtained is not None and not mark.is_absent:
                    grand_total_obtained += mark.marks_obtained
                grand_total_possible += es_item.total_marks
            overall_pct = float(grand_total_obtained / grand_total_possible * 100) if grand_total_possible > 0 else 0

        overall_grade = '-'
        for gs in grade_scales:
            if float(gs.min_percentage) <= overall_pct <= float(gs.max_percentage):
                overall_grade = gs.grade_label
                break

        # Build flattened subject-level summary for the frontend
        subject_summaries = []
        for subj_id, subj_name in all_subjects.items():
            subj_total = Decimal('0')
            subj_obtained = Decimal('0')
            subj_absent = False
            subj_pass = True

            for exam in exams:
                for es_item in es_by_exam.get(exam.id, []):
                    if es_item.subject_id == subj_id:
                        mark = marks_lookup.get(es_item.id)
                        subj_total += es_item.total_marks
                        if mark and mark.marks_obtained is not None and not mark.is_absent:
                            subj_obtained += mark.marks_obtained
                            if mark.marks_obtained < es_item.passing_marks:
                                subj_pass = False
                        else:
                            subj_pass = False
                            if mark and mark.is_absent:
                                subj_absent = True

            subj_pct = float(subj_obtained / subj_total * 100) if subj_total > 0 else 0
            subj_grade = '-'
            for gs in grade_scales:
                if float(gs.min_percentage) <= subj_pct <= float(gs.max_percentage):
                    subj_grade = gs.grade_label
                    break

            subject_summaries.append({
                'subject_name': subj_name,
                'total_marks': float(subj_total),
                'marks_obtained': float(subj_obtained),
                'percentage': round(subj_pct, 2),
                'grade': subj_grade,
                'is_pass': subj_pass,
                'is_absent': subj_absent,
            })

        from academic_sessions.utils import resolve_class_display_name, resolve_current_academic_year_id

        enrollment_class_name = (
            enrollment.session_class.display_name
            if enrollment.session_class_id and enrollment.session_class.display_name
            else enrollment.class_obj.name
        )
        current_class_name = resolve_class_display_name(
            school_id, resolve_current_academic_year_id(school_id), student.class_obj,
        )

        return Response({
            'student_name': student.name,
            'roll_number': enrollment.roll_number or student.roll_number,
            'class_name': enrollment_class_name,
            'school_name': student.school.name,
            'academic_year_name': enrollment.academic_year.name,
            'term_name': exams[0].term.name if exams and exams[0].term else None,
            'enrollment_info': {
                'enrollment_id': enrollment.id,
                'class_at_report_session': enrollment_class_name,
                'current_class': current_class_name,
                'academic_year_id': enrollment.academic_year_id,
                'academic_year_name': enrollment.academic_year.name,
            },
            'student': {
                'id': student.id,
                'name': student.name,
                'roll_number': enrollment.roll_number or student.roll_number,
                'class_name': enrollment_class_name,
                'school_name': student.school.name,
            },
            'subjects': subject_summaries,
            'exams': exam_data,
            'summary': {
                'total_marks': float(grand_total_possible),
                'obtained_marks': float(grand_total_obtained),
                'total_obtained': float(grand_total_obtained),
                'total_possible': float(grand_total_possible),
                'percentage': round(overall_pct, 2),
                'grade': overall_grade,
                'overall_pass': all(s['is_pass'] for s in subject_summaries) if subject_summaries else False,
                'calculation_mode': 'weighted' if use_weighted and exams.count() > 1 else 'simple',
            },
            'grade_scales': [
                {
                    'grade_label': gs.grade_label,
                    'min_percentage': float(gs.min_percentage),
                    'max_percentage': float(gs.max_percentage),
                    'gpa_points': float(gs.gpa_points),
                }
                for gs in grade_scales
            ],
        })


# ===========================================
# Question Paper Builder ViewSets
# ===========================================


class QuestionViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    """ViewSet for Question management."""
    required_module = 'examinations'
    queryset = Question.objects.all()
    permission_classes = [IsAuthenticated, HasSchoolAccess]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return QuestionCreateUpdateSerializer
        return QuestionSerializer

    def get_queryset(self):
        qs = super().get_queryset()

        scope = resolve_class_scope(
            self.request,
            school_id=_resolve_school_id(self.request),
            class_param_names=('class_obj', 'class_id'),
        )
        if scope['invalid']:
            return qs.none()

        class_id = scope['class_obj_id']
        chapter_id = self.request.query_params.get('chapter_id')
        book_id = self.request.query_params.get('book_id')

        if get_effective_role(self.request) == 'TEACHER':
            school_id = _resolve_school_id(self.request)
            class_subject_map = _get_teacher_class_subject_map(self.request, school_id=school_id)
            if class_id:
                allowed_subject_ids = class_subject_map.get(class_id, set())
            else:
                allowed_subject_ids = _get_teacher_allowed_subject_ids(self.request, school_id=school_id)
            if not allowed_subject_ids:
                return qs.none()
            qs = qs.filter(subject_id__in=allowed_subject_ids)

            if class_id and class_id not in class_subject_map:
                return qs.none()

        if class_id:
            qs = qs.filter(tested_topics__chapter__book__class_obj_id=class_id)
        
        # Filter by subject
        subject_id = self.request.query_params.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)

        if book_id:
            qs = qs.filter(tested_topics__chapter__book_id=book_id)

        if chapter_id:
            qs = qs.filter(tested_topics__chapter_id=chapter_id)
        
        # Filter by exam type
        exam_type_id = self.request.query_params.get('exam_type')
        if exam_type_id:
            qs = qs.filter(exam_type_id=exam_type_id)
        
        # Filter by question type
        question_type = self.request.query_params.get('question_type')
        if question_type:
            qs = qs.filter(question_type=question_type)
        
        # Filter by difficulty
        difficulty = self.request.query_params.get('difficulty') or self.request.query_params.get('difficulty_level')
        if difficulty:
            qs = qs.filter(difficulty_level=difficulty)

        bloom_level = self.request.query_params.get('bloom_level')
        if bloom_level:
            qs = qs.filter(bloom_level=bloom_level)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        else:
            qs = qs.filter(is_active=True)
        
        # Search by question text
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(question_text__icontains=search)

        topic_id = self.request.query_params.get('topic_id')
        if topic_id:
            qs = qs.filter(tested_topics__id=topic_id).distinct()

        topic_ids = self.request.query_params.getlist('topics')
        if topic_ids:
            qs = qs.filter(tested_topics__id__in=topic_ids).distinct()

        tag_id = self.request.query_params.get('tag_id')
        if tag_id:
            qs = qs.filter(question_tags__tag_id=tag_id).distinct()

        ordering = self.request.query_params.get('ordering')
        if ordering in {'paper_use_count', '-paper_use_count'}:
            qs = qs.order_by(ordering, 'id')

        if class_id or book_id or chapter_id:
            qs = qs.distinct()

        return qs.select_related('subject', 'exam_type', 'created_by', 'stats')

    def perform_create(self, serializer):
        school_id = _resolve_school_id(self.request)
        if get_effective_role(self.request) == 'TEACHER':
            subject = serializer.validated_data.get('subject')
            allowed_subject_ids = _get_teacher_allowed_subject_ids(self.request, school_id=school_id)
            if not subject or subject.id not in allowed_subject_ids:
                raise PermissionDenied('You can only create questions for your assigned subjects.')
        serializer.save(school_id=school_id, created_by=self.request.user)

    def perform_update(self, serializer):
        school_id = _resolve_school_id(self.request)
        if get_effective_role(self.request) == 'TEACHER':
            subject = serializer.validated_data.get('subject', serializer.instance.subject)
            allowed_subject_ids = _get_teacher_allowed_subject_ids(self.request, school_id=school_id)
            if not subject or subject.id not in allowed_subject_ids:
                raise PermissionDenied('You can only edit questions for your assigned subjects.')
        serializer.instance._revision_changed_by = self.request.user
        serializer.save()

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(detail=True, methods=['post'], url_path='add_tag')
    def add_tag(self, request, pk=None):
        question = self.get_object()
        tag_id = request.data.get('tag_id')
        if not tag_id:
            return Response({'detail': 'tag_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            tag = Tag.objects.get(id=tag_id)
        except Tag.DoesNotExist:
            return Response({'detail': 'Tag not found.'}, status=status.HTTP_404_NOT_FOUND)

        if tag.school_id and tag.school_id != question.school_id:
            return Response({'detail': 'Tag does not belong to the same school.'}, status=status.HTTP_400_BAD_REQUEST)

        if request.data.get('remove'):
            deleted, _ = QuestionTag.objects.filter(question=question, tag=tag).delete()
            return Response({'removed': bool(deleted)})

        relation, created = QuestionTag.objects.get_or_create(question=question, tag=tag)
        return Response({'created': created, 'id': relation.id}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='semantic_search')
    def semantic_search(self, request):
        query = (request.query_params.get('q') or '').strip()
        if not query:
            return Response([])

        try:
            limit = max(1, min(int(request.query_params.get('limit', 10)), 50))
        except (TypeError, ValueError):
            limit = 10

        queryset = self.get_queryset().filter(embedding__isnull=False)
        if not queryset.exists():
            return Response([])

        query_embedding = generate_text_embedding(query)
        matches = list(
            queryset.prefetch_related('tested_topics__chapter__book')
            .annotate(similarity_distance=CosineDistance('embedding', query_embedding))
            .order_by('similarity_distance')[:limit]
        )

        results = []
        for question in matches:
            first_topic = next(iter(question.tested_topics.all()), None)
            chapter = first_topic.chapter if first_topic else None
            book = chapter.book if chapter else None
            results.append({
                'id': question.id,
                'question_text': question.question_text,
                'question_type': question.question_type,
                'difficulty_level': question.difficulty_level,
                'marks': str(question.marks),
                'similarity_score': max(0.0, 1.0 - float(question.similarity_distance)),
                'chapter_title': chapter.title if chapter else '',
                'topic_title': first_topic.title if first_topic else '',
                'book_title': book.title if book else '',
            })

        return Response(results)
    
    @action(detail=False, methods=['post'])
    def generate_from_lesson(self, request):
        """
        Generate AI questions from a lesson plan.
        
        Body: {
            lesson_plan_id: int,
            question_count: int (5-20),
            question_type: str (MCQ/SHORT/ESSAY/TRUE_FALSE),
            difficulty_level: str (EASY/MEDIUM/HARD)
        }
        
        Returns: {questions: [...], message: "..."}
        """
        from django.conf import settings
        from rest_framework import status
        from lms.models import LessonPlan
        import requests
        import json
        import re
        
        lesson_plan_id = request.data.get('lesson_plan_id')
        question_count = request.data.get('question_count', 5)
        question_type = request.data.get('question_type', 'MCQ')
        difficulty_level = request.data.get('difficulty_level', 'MEDIUM')
        
        # Validate inputs
        if not lesson_plan_id:
            return Response(
                {'error': 'lesson_plan_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not (5 <= question_count <= 20):
            return Response(
                {'error': 'question_count must be between 5 and 20'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Fetch lesson plan
        try:
            lesson = LessonPlan.objects.get(
                id=lesson_plan_id,
                school=request.tenant_school
            )
        except LessonPlan.DoesNotExist:
            return Response(
                {'error': 'Lesson plan not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        school_id = _resolve_school_id(request)
        if not _is_teacher_allowed_for_class_subject(
            request,
            lesson.class_obj_id,
            lesson.subject_id,
            school_id=school_id,
        ):
            return Response(
                {'error': 'You can only generate questions for your assigned class-subjects.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        # Get topics
        topics = lesson.planned_topics.select_related(
            'chapter', 'chapter__book'
        ).all()
        
        if not topics:
            return Response(
                {'error': 'Lesson plan has no topics selected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Build AI prompt
        topics_text = '\n'.join([
            f"- Chapter {t.chapter.chapter_number}: {t.chapter.title}\n"
            f"  Topic {t.topic_number}: {t.title}\n"
            f"  Description: {t.description or 'N/A'}"
            for t in topics
        ])
        
        prompt = f"""You are an expert educator creating {question_type} questions for {lesson.subject.name} exam at {lesson.class_obj.name} level, {difficulty_level.lower()} difficulty.

Generate exactly {question_count} questions based on these topics:

{topics_text}

For each question:
1. Write clear, concise question text
2. For MCQ: provide 4 options (A, B, C, D) with one correct answer
3. Specify which topic (e.g., "3.2") it tests
4. Assign marks

Respond with ONLY a JSON array, no extra text:
[
  {{
    "question_text": "...",
    "question_type": "{question_type}",
    "options": {{"A": "...", "B": "...", "C": "...", "D": "..."}},
    "correct_answer": "A",
    "tested_topic_number": "3.2",
    "marks": 2
  }}
]
"""

        ai_job = create_ai_job(
            job_type='generate_questions',
            triggered_by=request.user,
            school=lesson.school,
            input_data={
                'lesson_plan_id': lesson_plan_id,
                'question_count': question_count,
                'question_type': question_type,
                'difficulty_level': difficulty_level,
            },
            model_used=getattr(settings, 'GROQ_MODEL', 'unknown'),
        )
        
        # Call Groq API
        try:
            groq_response = requests.post(
                'https://api.groq.com/openai/v1/chat/completions',
                headers={
                    'Authorization': f'Bearer {settings.GROQ_API_KEY}',
                    'Content-Type': 'application/json',
                },
                json={
                    'model': settings.GROQ_MODEL,
                    'messages': [{'role': 'user', 'content': prompt}],
                    'temperature': 0.7,
                    'max_tokens': 2048,
                },
                timeout=30,
            )
            groq_response.raise_for_status()
            
            # Parse response
            ai_text = groq_response.json()['choices'][0]['message']['content'].strip()
            
            # Extract JSON from response
            json_match = re.search(r'\[.*\]', ai_text, re.DOTALL)
            if json_match:
                questions_data = json.loads(json_match.group())
            else:
                questions_data = json.loads(ai_text)
            
            # Create Question objects
            created_questions = []
            for q_data in questions_data:
                # Parse topic number "3.2"
                topic_num_str = q_data.get('tested_topic_number', '')
                parts = topic_num_str.split('.')
                tested_topic = None
                
                if len(parts) == 2:
                    try:
                        ch_num, t_num = int(parts[0]), int(parts[1])
                        for t in topics:
                            if (t.chapter.chapter_number == ch_num and 
                                t.topic_number == t_num):
                                tested_topic = t
                                break
                    except ValueError:
                        pass
                
                # Create question
                question = Question.objects.create(
                    school=request.tenant_school,
                    subject=lesson.subject,
                    question_text=q_data.get('question_text', ''),
                    question_type=question_type,
                    difficulty_level=difficulty_level,
                    marks=q_data.get('marks', 1),
                    option_a=q_data.get('options', {}).get('A', ''),
                    option_b=q_data.get('options', {}).get('B', ''),
                    option_c=q_data.get('options', {}).get('C', ''),
                    option_d=q_data.get('options', {}).get('D', ''),
                    correct_answer=q_data.get('correct_answer', ''),
                    created_by=request.user,
                )
                
                # Link to topic
                if tested_topic:
                    question.tested_topics.add(tested_topic)
                
                created_questions.append(question)
            
            serializer = QuestionSerializer(created_questions, many=True)
            complete_ai_job(
                ai_job,
                output_data={
                    'question_ids': [question.id for question in created_questions],
                    'question_count': len(created_questions),
                },
                accepted=True,
            )
            return Response({
                'message': f'Generated {len(created_questions)} questions',
                'ai_job_id': ai_job.id,
                'questions': serializer.data,
            }, status=status.HTTP_201_CREATED)
            
        except requests.RequestException as e:
            fail_ai_job(ai_job, error_message=e)
            return Response(
                {'error': f'API error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except json.JSONDecodeError as e:
            fail_ai_job(ai_job, error_message=e)
            return Response(
                {'error': f'Invalid JSON from AI: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            fail_ai_job(ai_job, error_message=e)
            return Response(
                {'error': f'Generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def by_lesson_plan(self, request):
        """
        Get all questions for a lesson plan's topics.
        Query params: lesson_plan_id (required)
        """
        from lms.models import LessonPlan
        
        lesson_plan_id = request.query_params.get('lesson_plan_id')
        if not lesson_plan_id:
            return Response(
                {'error': 'lesson_plan_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            lesson = LessonPlan.objects.get(
                id=lesson_plan_id,
                school=request.tenant_school
            )
        except LessonPlan.DoesNotExist:
            return Response(
                {'error': 'Lesson plan not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        school_id = _resolve_school_id(request)
        if not _is_teacher_allowed_for_class_subject(
            request,
            lesson.class_obj_id,
            lesson.subject_id,
            school_id=school_id,
        ):
            return Response(
                {'error': 'You can only access questions for your assigned class-subjects.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        topic_ids = lesson.planned_topics.values_list('id', flat=True)
        qs = self.get_queryset().filter(tested_topics__id__in=topic_ids).distinct()
        
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)


class ExamPaperViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    """ViewSet for ExamPaper management."""
    required_module = 'examinations'
    queryset = ExamPaper.objects.all()
    permission_classes = [IsAuthenticated, HasSchoolAccess]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ExamPaperCreateUpdateSerializer
        return ExamPaperSerializer

    def get_queryset(self):
        qs = super().get_queryset()

        if get_effective_role(self.request) == 'TEACHER':
            school_id = _resolve_school_id(self.request)
            class_subject_map = _get_teacher_class_subject_map(self.request, school_id=school_id)
            predicates = Q()
            for class_id, subject_ids in class_subject_map.items():
                if subject_ids:
                    predicates |= Q(class_obj_id=class_id, subject_id__in=list(subject_ids))
            if not predicates:
                return qs.none()
            qs = qs.filter(predicates)
        
        # Filter by class
        class_id = self.request.query_params.get('class_obj')
        if class_id:
            qs = qs.filter(class_obj_id=class_id)
        
        # Filter by subject
        subject_id = self.request.query_params.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        
        # Filter by exam
        exam_id = self.request.query_params.get('exam')
        if exam_id:
            qs = qs.filter(exam_id=exam_id)
        
        # Filter by status
        paper_status = self.request.query_params.get('status')
        if paper_status:
            qs = qs.filter(status=paper_status)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        else:
            qs = qs.filter(is_active=True)
        
        # Search by title
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(paper_title__icontains=search)
        
        return qs.select_related(
            'class_obj', 'subject', 'exam', 'exam_subject', 'generated_by'
        ).prefetch_related('paper_questions__question')

    def perform_create(self, serializer):
        school_id = _resolve_school_id(self.request)
        class_obj = serializer.validated_data.get('class_obj')
        subject = serializer.validated_data.get('subject')
        if not _can_manage_exam_papers(
            self.request,
            class_id=getattr(class_obj, 'id', None),
            subject_id=getattr(subject, 'id', None),
            school_id=school_id,
        ):
            raise PermissionDenied('Only School Admin, Principal, or assigned class teachers can create exam papers.')
        serializer.save(school_id=school_id, generated_by=self.request.user)

    def perform_update(self, serializer):
        school_id = _resolve_school_id(self.request)
        class_obj = serializer.validated_data.get('class_obj', serializer.instance.class_obj)
        subject = serializer.validated_data.get('subject', serializer.instance.subject)
        if not _can_manage_exam_papers(
            self.request,
            class_id=getattr(class_obj, 'id', None),
            subject_id=getattr(subject, 'id', None),
            school_id=school_id,
        ):
            raise PermissionDenied('Only School Admin, Principal, or assigned class teachers can edit exam papers.')
        serializer.save()

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    def _validate_paper_manage_scope(self, class_obj, subject):
        school_id = _resolve_school_id(self.request)
        if not _can_manage_exam_papers(
            self.request,
            class_id=getattr(class_obj, 'id', None),
            subject_id=getattr(subject, 'id', None),
            school_id=school_id,
        ):
            raise PermissionDenied('Only School Admin, Principal, or assigned class teachers can create or edit exam papers.')
        return school_id

    def _save_manual_draft_questions(self, exam_paper, manual_questions):
        existing_assignments = {
            paper_question.question_id: paper_question
            for paper_question in exam_paper.paper_questions.select_related('question').all()
        }
        retained_assignment_ids = set()

        for index, raw_question in enumerate(manual_questions, start=1):
            question_payload = dict(raw_question)
            question_id = question_payload.pop('question_id', None)
            question_order = question_payload.pop('question_order', index)
            marks_override = question_payload.pop('marks_override', None)
            section_key = question_payload.pop('section_key', '') or ''
            question_payload.pop('local_id', None)

            assignment = None
            question_instance = None
            if question_id is not None:
                assignment = existing_assignments.get(question_id)
                if assignment is not None:
                    question_instance = assignment.question
                else:
                    # Not yet linked to this paper (e.g. just attached from the question
                    # bank picker) — reuse the existing bank question instead of raising or
                    # creating a duplicate Question row.
                    question_instance = Question.objects.filter(
                        id=question_id, school=exam_paper.school,
                    ).first()
                    if question_instance is None:
                        raise ValidationError({
                            'manual_questions': [f'question_id {question_id} was not found.']
                        })

            question_payload['subject'] = exam_paper.subject_id
            serializer = QuestionCreateUpdateSerializer(
                instance=question_instance,
                data=question_payload,
                partial=question_instance is not None,
                context={'request': self.request},
            )
            serializer.is_valid(raise_exception=True)

            if question_instance is None:
                question = serializer.save(
                    school=exam_paper.school,
                    created_by=self.request.user,
                )
            else:
                question = serializer.save()

            if assignment is None:
                assignment = PaperQuestion(
                    exam_paper=exam_paper,
                    question=question,
                )

            assignment.question_order = question_order
            assignment.section_key = str(section_key)[:50]
            assignment.marks_override = marks_override
            assignment.save()
            assignment.sync_question_snapshot()
            retained_assignment_ids.add(assignment.id)

        exam_paper.paper_questions.exclude(id__in=retained_assignment_ids).delete()

    @action(detail=False, methods=['post'], url_path='ensure-draft')
    def ensure_draft(self, request):
        """Create or refresh a server-backed draft paper before autosave begins."""
        draft_id = request.data.get('draft_id') or request.data.get('id')
        exam_paper = None

        if draft_id:
            exam_paper = self.get_queryset().filter(pk=draft_id).first()
            if exam_paper is None:
                return Response({'detail': 'Draft paper not found.'}, status=status.HTTP_404_NOT_FOUND)
            if exam_paper.status != ExamPaper.Status.DRAFT:
                return Response(
                    {'detail': 'Only draft papers can be resumed for autosave.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        serializer = ExamPaperDraftEnsureSerializer(
            instance=exam_paper,
            data=request.data,
            partial=exam_paper is not None,
        )
        serializer.is_valid(raise_exception=True)

        class_obj = serializer.validated_data.get('class_obj', getattr(exam_paper, 'class_obj', None))
        subject = serializer.validated_data.get('subject', getattr(exam_paper, 'subject', None))
        school_id = self._validate_paper_manage_scope(class_obj, subject)

        with transaction.atomic():
            if exam_paper is None:
                exam_paper = serializer.save(
                    school_id=school_id,
                    generated_by=request.user,
                    status=ExamPaper.Status.DRAFT,
                )
                http_status = status.HTTP_201_CREATED
            else:
                exam_paper = serializer.save(status=ExamPaper.Status.DRAFT)
                http_status = status.HTTP_200_OK

        return Response(ExamPaperSerializer(exam_paper).data, status=http_status)

    @action(detail=True, methods=['post'])
    def autosave(self, request, pk=None):
        """Autosave draft metadata and manual-entry questions into the question bank."""
        exam_paper = self.get_object()
        if exam_paper.status != ExamPaper.Status.DRAFT:
            return Response(
                {'detail': 'Only draft papers can be autosaved.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = ExamPaperDraftAutosaveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated = serializer.validated_data
        manual_questions = validated.pop('manual_questions', None)

        class_obj = validated.get('class_obj', exam_paper.class_obj)
        subject = validated.get('subject', exam_paper.subject)
        self._validate_paper_manage_scope(class_obj, subject)

        with transaction.atomic():
            for attr, value in validated.items():
                setattr(exam_paper, attr, value)
            exam_paper.save()

            if manual_questions is not None:
                self._save_manual_draft_questions(exam_paper, manual_questions)

        exam_paper.refresh_from_db()
        return Response(ExamPaperSerializer(exam_paper).data)

    @action(detail=True, methods=['get'], url_path='generate-pdf')
    def generate_pdf(self, request, pk=None):
        """Generate and download PDF for this exam paper."""
        from .pdf_generator import ExamPaperPDFGenerator
        
        exam_paper = self.get_object()
        
        try:
            generator = ExamPaperPDFGenerator(exam_paper)
            pdf_bytes = generator.generate()
            
            # Create filename
            filename = f"{exam_paper.paper_title.replace(' ', '_')}.pdf"
            
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
        
        except Exception as e:
            return Response(
                {'detail': f'Error generating PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='generate-docx')
    def generate_docx(self, request, pk=None):
        """Generate and download DOCX for this exam paper."""
        from .docx_generator import ExamPaperDOCXGenerator

        exam_paper = self.get_object()

        try:
            generator = ExamPaperDOCXGenerator(exam_paper)
            docx_bytes = generator.generate()

            filename = f"{exam_paper.paper_title.replace(' ', '_')}.docx"

            response = HttpResponse(
                docx_bytes,
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {'detail': f'Error generating DOCX: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def link_lesson_plans(self, request, pk=None):
        """
        Link lesson plans to this exam paper.
        Body: {lesson_plan_ids: [1, 2, 3]}
        """
        from lms.models import LessonPlan
        
        exam_paper = self.get_object()
        lesson_plan_ids = request.data.get('lesson_plan_ids', [])
        
        lesson_plans = LessonPlan.objects.filter(
            id__in=lesson_plan_ids,
            school=request.tenant_school
        )
        
        exam_paper.lesson_plans.set(lesson_plans)
        
        serializer = self.get_serializer(exam_paper)
        return Response({
            'message': f'Linked {lesson_plans.count()} lesson plans',
            'exam_paper': serializer.data
        })
    
    @action(detail=True, methods=['get'])
    def coverage_stats(self, request, pk=None):
        """
        Get coverage statistics for this exam paper.
        Returns: topics count, covered topics, lesson plans, etc.
        """
        exam_paper = self.get_object()
        slo_coverage_count = exam_paper.covered_topics.filter(
            standard_alignments__isnull=False,
        ).values('standard_alignments__objective_id').distinct().count()
        
        return Response({
            'exam_paper_id': exam_paper.id,
            'paper_title': exam_paper.paper_title,
            'total_questions': exam_paper.question_count,
            'total_marks': exam_paper.total_marks,
            'covered_topics': [
                {
                    'id': t.id,
                    'chapter': f"{t.chapter.chapter_number}: {t.chapter.title}",
                    'topic': f"{t.topic_number}: {t.title}",
                    'questions_count': t.test_questions.filter(
                        paper_assignments__exam_paper=exam_paper
                    ).count(),
                }
                for t in exam_paper.covered_topics
            ],
            'linked_lesson_plans': [
                {
                    'id': lp.id,
                    'title': lp.title,
                    'lesson_date': lp.lesson_date,
                }
                for lp in exam_paper.lesson_plans.all()
            ],
            'topic_count': exam_paper.covered_topics.count(),
            'slo_coverage_count': slo_coverage_count,
            # Backward-compatible aliases used by older clients/tests.
            'covered_slos': slo_coverage_count,
            'slo_coverage': slo_coverage_count,
        })
    
    @action(detail=False, methods=['post'])
    def create_from_lessons(self, request):
        """
        Create exam paper from lesson plans.
        
        Body: {
            lesson_plan_ids: [1, 2, 3],
            class_id: 5,
            subject_id: 10,
            paper_title: "Mid-Term Exam",
            instructions: "...",
            total_marks: 100,
            duration_minutes: 60,
            question_type: "MCQ",
            difficulty_balance: {"EASY": 0.3, "MEDIUM": 0.5, "HARD": 0.2}
        }
        """
        from lms.models import LessonPlan
        
        lesson_plan_ids = request.data.get('lesson_plan_ids', [])
        class_id = request.data.get('class_id')
        subject_id = request.data.get('subject_id')
        paper_title = request.data.get('paper_title')
        instructions = request.data.get('instructions', '')
        total_marks = request.data.get('total_marks', 100)
        duration_minutes = request.data.get('duration_minutes', 60)
        
        if not (lesson_plan_ids and class_id and subject_id and paper_title):
            return Response(
                {'error': 'Missing required fields'},
                status=status.HTTP_400_BAD_REQUEST
            )

        school_id = _resolve_school_id(request)
        if not _can_manage_exam_papers(
            request,
            class_id=class_id,
            subject_id=subject_id,
            school_id=school_id,
        ):
            return Response(
                {'error': 'Only School Admin, Principal, or assigned class teachers can create exam papers.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        # Fetch lesson plans
        lesson_plans = LessonPlan.objects.filter(
            id__in=lesson_plan_ids,
            school=request.tenant_school
        )
        
        if not lesson_plans.exists():
            return Response(
                {'error': 'No lesson plans found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get topics from lesson plans
        topic_ids = set()
        for lp in lesson_plans:
            topic_ids.update(lp.planned_topics.values_list('id', flat=True))
        
        if not topic_ids:
            return Response(
                {'error': 'Selected lesson plans have no topics'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get questions for those topics
        questions_qs = Question.objects.filter(
            school=request.tenant_school,
            subject_id=subject_id,
            tested_topics__id__in=topic_ids,
            is_active=True
        ).distinct()
        
        if not questions_qs.exists():
            return Response(
                {'error': 'No questions available for selected topics'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create exam paper
        exam_paper = ExamPaper.objects.create(
            school=request.tenant_school,
            class_obj_id=class_id,
            subject_id=subject_id,
            paper_title=paper_title,
            instructions=instructions,
            total_marks=total_marks,
            duration_minutes=duration_minutes,
            status='DRAFT',
            generated_by=request.user,
        )
        
        # Link lesson plans
        exam_paper.lesson_plans.set(lesson_plans)
        
        # Add questions (balance by difficulty if needed)
        selected_questions = list(questions_qs[:15])  # Default: up to 15 questions
        
        for idx, q in enumerate(selected_questions):
            paper_question = PaperQuestion.objects.create(
                exam_paper=exam_paper,
                question=q,
                question_order=idx + 1,
                marks_override=q.marks,
            )
            paper_question.sync_question_snapshot()
        
        serializer = ExamPaperSerializer(exam_paper)
        return Response({
            'message': f'Created paper with {len(selected_questions)} questions',
            'exam_paper': serializer.data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='review-questions')
    def review_questions(self, request):
        """AI-powered grammar and spelling review for questions."""
        from .paper_ocr_processor import QuestionReviewAI
        
        serializer = QuestionReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        question_texts = serializer.validated_data['questions']
        
        try:
            reviewer = QuestionReviewAI()
            results = reviewer.review_questions(question_texts)
            
            return Response({'results': results}, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {'detail': f'Error reviewing questions: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PaperUploadViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ModelViewSet):
    """ViewSet for PaperUpload management (image uploads for OCR)."""
    required_module = 'examinations'
    queryset = PaperUpload.objects.all()
    permission_classes = [IsAuthenticated, HasSchoolAccess]
    serializer_class = PaperUploadSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        
        # Filter by status
        upload_status = self.request.query_params.get('status')
        if upload_status:
            qs = qs.filter(status=upload_status)
        
        # Filter by uploaded user
        if self.request.query_params.get('my_uploads') == 'true':
            qs = qs.filter(uploaded_by=self.request.user)
        
        return qs.select_related('school', 'exam_paper', 'uploaded_by').order_by('-created_at')

    @action(detail=False, methods=['post'], url_path='upload-image')
    def upload_image(self, request):
        """Upload paper image and trigger OCR processing."""
        from core.storage import SupabaseStorageService
        from .tasks import process_paper_upload_ocr
        
        serializer = PaperUploadCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        image_file = serializer.validated_data['image']
        context_class_id = serializer.validated_data.get('class_obj')
        context_subject_id = serializer.validated_data.get('subject')
        school_id = _resolve_school_id(request)
        
        if not school_id:
            return Response(
                {'detail': 'School ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Upload to Supabase storage
            storage_service = SupabaseStorageService()
            
            # Use a folder structure: papers/{school_id}/{timestamp}
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            folder_path = f"papers/{school_id}"
            
            image_url = storage_service.upload_file(
                file=image_file,
                folder= folder_path,
                filename=f"paper_{timestamp}_{image_file.name}"
            )
            
            # Create PaperUpload record
            upload = PaperUpload.objects.create(
                school_id=school_id,
                uploaded_by=request.user,
                image_url=image_url,
                context_class_id=context_class_id,
                context_subject_id=context_subject_id,
                status=PaperUpload.Status.PENDING
            )
            
            # Trigger async OCR processing
            from core.task_utils import call_task
            call_task(process_paper_upload_ocr, upload.id)
            
            return Response(
                PaperUploadSerializer(upload).data,
                status=status.HTTP_201_CREATED
            )
        
        except Exception as e:
            return Response(
                {'detail': f'Error uploading image: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='confirm')
    def confirm_extraction(self, request, pk=None):
        """Confirm extracted questions.

        Two modes:
        - exam_paper_id provided (current draft-pipeline flow): the paper and its
          questions were already created via ensure-draft + autosave, using the same
          manual_questions path as typed entry. This call only records the
          PaperFeedback learning-loop row and marks/links the upload — it never
          creates ExamPaper/Question/PaperQuestion rows itself.
        - exam_paper_id absent (legacy one-shot flow, kept for API compatibility):
          creates the ExamPaper and its Questions directly from confirmed_data.
        """
        upload = self.get_object()

        if upload.status != PaperUpload.Status.EXTRACTED:
            return Response(
                {'detail': 'Upload must be in EXTRACTED status'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get confirmed data from request
        confirmed_json = request.data.get('confirmed_data')
        paper_metadata = request.data.get('paper_metadata', {})
        exam_paper_id = request.data.get('exam_paper_id')

        if not confirmed_json:
            return Response(
                {'detail': 'confirmed_data is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            school_id = _resolve_school_id(request)

            if exam_paper_id:
                exam_paper = ExamPaper.objects.filter(id=exam_paper_id, school_id=school_id).first()
                if exam_paper is None:
                    return Response(
                        {'detail': 'exam_paper_id was not found for this school.'},
                        status=status.HTTP_404_NOT_FOUND,
                    )

                if not _can_manage_exam_papers(
                    request,
                    class_id=exam_paper.class_obj_id,
                    subject_id=exam_paper.subject_id,
                    school_id=school_id,
                ):
                    return Response(
                        {'detail': 'Only School Admin, Principal, or assigned class teachers can confirm exam papers.'},
                        status=status.HTTP_403_FORBIDDEN,
                    )

                questions = confirmed_json.get('questions', [])

                # Feedback-only: the draft pipeline already created the paper/questions.
                PaperFeedback.objects.create(
                    paper_upload=upload,
                    ai_extracted_json=upload.ai_extracted_json,
                    user_confirmed_json=confirmed_json,
                    accuracy_metrics={
                        'total_questions': len(questions),
                        'extraction_confidence': upload.extraction_confidence
                    },
                    confirmed_by=request.user
                )

                upload.status = PaperUpload.Status.CONFIRMED
                upload.exam_paper = exam_paper
                upload.save()

                return Response(
                    {
                        'detail': 'Paper successfully confirmed',
                        'exam_paper_id': exam_paper.id,
                        'questions_created': exam_paper.question_count,
                    },
                    status=status.HTTP_200_OK
                )

            # Legacy one-shot flow: no exam_paper_id, so create everything here.
            if not _can_manage_exam_papers(
                request,
                class_id=paper_metadata.get('class_obj'),
                subject_id=paper_metadata.get('subject'),
                school_id=school_id,
            ):
                return Response(
                    {'detail': 'Only School Admin, Principal, or assigned class teachers can create exam papers.'},
                    status=status.HTTP_403_FORBIDDEN,
                )

            # Create ExamPaper
            exam_paper = ExamPaper.objects.create(
                school_id=school_id,
                class_obj_id=paper_metadata.get('class_obj'),
                subject_id=paper_metadata.get('subject'),
                exam_id=paper_metadata.get('exam'),
                exam_subject_id=paper_metadata.get('exam_subject'),
                paper_title=paper_metadata.get('paper_title', 'Untitled Paper'),
                instructions=paper_metadata.get('instructions', ''),
                total_marks=paper_metadata.get('total_marks', 100),
                duration_minutes=paper_metadata.get('duration_minutes', 60),
                status=ExamPaper.Status.DRAFT,
                generated_by=request.user
            )

            # Create Questions from confirmed data
            questions = confirmed_json.get('questions', [])
            for idx, q_data in enumerate(questions, start=1):
                question = Question.objects.create(
                    school_id=school_id,
                    subject_id=paper_metadata.get('subject'),
                    question_text=q_data.get('question_text', ''),
                    question_type=q_data.get('question_type', 'SHORT'),
                    difficulty_level=q_data.get('difficulty_level', 'MEDIUM'),
                    marks=q_data.get('marks', 1),
                    option_a=q_data.get('options', {}).get('A', ''),
                    option_b=q_data.get('options', {}).get('B', ''),
                    option_c=q_data.get('options', {}).get('C', ''),
                    option_d=q_data.get('options', {}).get('D', ''),
                    created_by=request.user,
                )

                # Link question to paper
                paper_question = PaperQuestion.objects.create(
                    exam_paper=exam_paper,
                    question=question,
                    question_order=idx,
                    marks_override=q_data.get('marks')
                )
                paper_question.sync_question_snapshot()

            # Create feedback record for learning loop
            PaperFeedback.objects.create(
                paper_upload=upload,
                ai_extracted_json=upload.ai_extracted_json,
                user_confirmed_json=confirmed_json,
                accuracy_metrics={
                    'total_questions': len(questions),
                    'extraction_confidence': upload.extraction_confidence
                },
                confirmed_by=request.user
            )

            # Update upload status
            upload.status = PaperUpload.Status.CONFIRMED
            upload.exam_paper = exam_paper
            upload.save()

            return Response(
                {
                    'detail': 'Paper successfully confirmed',
                    'exam_paper_id': exam_paper.id,
                    'questions_created': len(questions)
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {'detail': f'Error confirming extraction: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PaperFeedbackViewSet(ModuleAccessMixin, TenantQuerySetMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet for PaperFeedback (read-only for analytics)."""
    required_module = 'examinations'
    queryset = PaperFeedback.objects.all()
    permission_classes = [IsAuthenticated, IsSchoolAdmin, HasSchoolAccess]
    serializer_class = PaperFeedbackSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related('paper_upload', 'confirmed_by').order_by('-created_at')
