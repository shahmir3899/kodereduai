"""
Base report generator with PDF and Excel rendering capabilities.
Uses openpyxl for Excel and reportlab for PDF.
"""

import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class BaseReportGenerator:
    """
    Base class for all report generators.

    Subclasses implement:
        - get_data() -> dict of data for the report
        - render_pdf(data) -> bytes
        - render_excel(data) -> bytes
    """

    def __init__(self, school, parameters=None):
        self.school = school
        self.parameters = parameters or {}

    def _academic_year_id(self):
        academic_year_id = self.parameters.get('academic_year')
        if academic_year_id in (None, ''):
            return None
        try:
            return int(academic_year_id)
        except (TypeError, ValueError):
            return None

    def _session_class_id(self):
        """Section id from `parameters`, if the caller passed one.

        Filtering by class_obj_id alone pools every section sharing that master
        class together (e.g. "Class 2 - A" and "Class 2 - B") — callers that scope
        a report to a class should prefer this over class_obj_id when available.
        """
        session_class_id = self.parameters.get('session_class_id') or self.parameters.get('session_class')
        if session_class_id in (None, ''):
            return None
        try:
            return int(session_class_id)
        except (TypeError, ValueError):
            return None

    def _get_enrollment_map(self, student_ids):
        academic_year_id = self._academic_year_id()
        if not academic_year_id or not student_ids:
            return {}

        from academic_sessions.models import StudentEnrollment

        enrollment_qs = StudentEnrollment.objects.filter(
            school_id=self.school.id,
            academic_year_id=academic_year_id,
            is_active=True,
            student_id__in=student_ids,
        ).select_related('class_obj', 'session_class')
        return {enrollment.student_id: enrollment for enrollment in enrollment_qs}

    def _resolve_class_name(self, student, enrollment_map):
        enrollment = enrollment_map.get(student.id)
        if enrollment:
            if enrollment.session_class_id and enrollment.session_class:
                return enrollment.session_class.display_name
            if enrollment.class_obj_id and enrollment.class_obj:
                return enrollment.class_obj.name
        return student.class_obj.name if student.class_obj else 'Unknown'

    def _resolve_roll_number(self, student, enrollment_map):
        enrollment = enrollment_map.get(student.id)
        if enrollment and enrollment.roll_number:
            return enrollment.roll_number
        return student.roll_number

    def get_data(self) -> dict:
        raise NotImplementedError

    def render_pdf(self, data: dict) -> bytes:
        """
        Generate a PDF report.
        Uses reportlab if available, otherwise falls back to a simple text-based PDF.
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.units import inch

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=6,
            )

            # Student/subject photo, if provided, sits beside the title
            photo_bytes = data.get('photo_bytes')
            if photo_bytes:
                try:
                    photo = Image(io.BytesIO(photo_bytes), width=0.9*inch, height=0.9*inch)
                    title_block = [
                        [photo, Paragraph(f"{self.school.name}<br/>{data.get('title', 'Report')}", title_style)],
                    ]
                    photo_table = Table(title_block, colWidths=[1*inch, None])
                    photo_table.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ]))
                    elements.append(photo_table)
                except Exception:
                    logger.warning("Failed to embed photo in report", exc_info=True)
                    elements.append(Paragraph(self.school.name, title_style))
                    elements.append(Paragraph(data.get('title', 'Report'), styles['Heading2']))
            else:
                elements.append(Paragraph(self.school.name, title_style))
                elements.append(Paragraph(data.get('title', 'Report'), styles['Heading2']))
            elements.append(Spacer(1, 12))

            # Subtitle / date range
            if data.get('subtitle'):
                elements.append(Paragraph(data['subtitle'], styles['Normal']))
                elements.append(Spacer(1, 8))

            # Summary stats
            if data.get('summary'):
                for key, value in data['summary'].items():
                    elements.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
                elements.append(Spacer(1, 12))

            # Primary table (e.g. exam results)
            self._add_table_section(elements, styles, data.get('table_headers'), data.get('table_rows'))

            # Secondary table (e.g. LMS assignments)
            if data.get('lms_headers') and data.get('lms_rows'):
                elements.append(Spacer(1, 12))
                elements.append(Paragraph(data.get('lms_heading', 'Assignments'), styles['Heading3']))
                elements.append(Spacer(1, 6))
                self._add_table_section(elements, styles, data.get('lms_headers'), data.get('lms_rows'))

            # Footer
            elements.append(Spacer(1, 20))
            footer_text = f"Generated on {datetime.now().strftime('%d %B %Y at %I:%M %p')} | EducationAI"
            elements.append(Paragraph(footer_text, styles['Normal']))

            doc.build(elements)
            return buffer.getvalue()

        except ImportError:
            logger.warning("reportlab not installed, generating simple text PDF")
            return self._fallback_pdf(data)

    def _add_table_section(self, elements, styles, headers, rows):
        """Append a styled reportlab Table for (headers, rows) onto elements, if present."""
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle

        if not headers or not rows:
            return

        table_data = [headers] + rows
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

    def _fallback_pdf(self, data: dict) -> bytes:
        """Simple text-based fallback when reportlab is not available."""
        lines = [
            self.school.name,
            data.get('title', 'Report'),
            '=' * 50,
            '',
        ]
        if data.get('summary'):
            for k, v in data['summary'].items():
                lines.append(f"{k}: {v}")
            lines.append('')

        if data.get('table_headers') and data.get('table_rows'):
            lines.append('\t'.join(str(h) for h in data['table_headers']))
            lines.append('-' * 80)
            for row in data['table_rows']:
                lines.append('\t'.join(str(c) for c in row))

        lines.append('')
        lines.append(f"Generated: {datetime.now().strftime('%d %B %Y %I:%M %p')}")
        return '\n'.join(lines).encode('utf-8')

    def render_excel(self, data: dict) -> bytes:
        """Generate an Excel report using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = data.get('title', 'Report')[:31]

            # Header
            ws.merge_cells('A1:F1')
            ws['A1'] = self.school.name
            ws['A1'].font = Font(size=14, bold=True)

            ws.merge_cells('A2:F2')
            ws['A2'] = data.get('title', 'Report')
            ws['A2'].font = Font(size=12, bold=True)

            if data.get('subtitle'):
                ws.merge_cells('A3:F3')
                ws['A3'] = data['subtitle']

            start_row = 5

            # Summary
            if data.get('summary'):
                for i, (k, v) in enumerate(data['summary'].items()):
                    ws.cell(row=start_row + i, column=1, value=k).font = Font(bold=True)
                    ws.cell(row=start_row + i, column=2, value=str(v))
                start_row += len(data['summary']) + 1

            # Table
            if data.get('table_headers') and data.get('table_rows'):
                header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=10)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )

                for col_idx, header in enumerate(data['table_headers'], 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border

                for row_idx, row in enumerate(data['table_rows'], start_row + 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border

                # Auto-width columns
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        except ImportError:
            logger.error("openpyxl not installed")
            return b''

    def generate(self, format='PDF') -> bytes:
        """Generate report in specified format."""
        data = self.get_data()
        if format == 'XLSX':
            return self.render_excel(data)
        return self.render_pdf(data)
